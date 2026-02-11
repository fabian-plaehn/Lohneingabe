from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import calendar
from database import Database
from master_data import MasterDataDatabase
from utils import (
    get_days_of_krank,
    get_days_of_urlaub,
    get_fahrstunden_for_name,
    get_normal_hours_per_month,
    get_verpflegungsgeld_for_name,
    is_holiday,
    is_weekend,
    calculate_skug,
)
from utils import (
    get_hours_of_krank,
    get_hours_of_urlaub,
    get_days_of_feiertag,
    get_hours_of_feiertag,
)
from utils import get_skug_hours_for_name
from datatypes import WorkerTypes


def AddBorders(border_one: Border, border_two: Border) -> Border:
    sides = ["left", "right", "top", "bottom", "diagonal", "vertical", "horizontal"]
    border_kwargs = {}
    for side in sides:
        try:
            side_one = getattr(border_one, side)
        except AttributeError:
            side_one = None
        try:
            side_two = getattr(border_two, side)
        except AttributeError:
            side_two = None
        if side_one is None and side_two is None:
            continue

        if side_one is not None:
            border_kwargs[side] = side_one if side_one.style is not None else side_two
        else:
            border_kwargs[side] = side_two
    return Border(**border_kwargs)


def addLattice(min_row, max_row, min_col, max_col, ws: Workbook):
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = AddBorders(
                cell.border,
                Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                ),
            )


def set_create_border(min_row, max_row, min_col, max_col, side_style, ws: Workbook):
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if cell.row == min_row:
                cell.border = AddBorders(cell.border, Border(top=side_style))
            if cell.row == max_row:
                cell.border = AddBorders(cell.border, Border(bottom=side_style))
            if cell.column == min_col:
                cell.border = AddBorders(cell.border, Border(left=side_style))
            if cell.column == max_col:
                cell.border = AddBorders(cell.border, Border(right=side_style))


SKUG_COLOR = "92d050"
UNTER_8H_COLOR = "b8cce4"
AN_AB_COLOR = "ff0000"
FREE_DAY_COLOR = "ffc000"  # Orange color for weekends and holidays

# Define thick border style
thick_border = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick"),
)

summary_labels = [
    "Gesamtstunden",
    "Feiertag",
    "Urlaubsstunden",
    "Krankstunden",
    "SKUG",
    "Summe",
    "Mehr-/Minderstd",
    "V.-Zuschuss [€]",
]


def build_workbook_top_to_bottom(
    year: int, month: int, db: Database, master_db: MasterDataDatabase
):
    unique_names = master_db.get_all_names_list()

    all_persons = master_db.get_all_names()
    person_lookup = {p["name"]: p for p in all_persons}
    for name in unique_names:
        person_lookup[name]["arbeits_entries"] = db.get_arbeitsstunden_for_month(
            year, month, name
        )

    if not unique_names:
        print("No names found in entries")
        return None

    names_for_normal_table = [
        name for name in unique_names if not person_lookup[name]["extra_table"]
    ]
    names_for_extra_table = [
        name for name in unique_names if person_lookup[name]["extra_table"]
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"
    current_col = 1
    names_per_section = 99
    next_column = 1

    num_sections = (
        len(names_for_normal_table) + names_per_section - 1
    ) // names_per_section

    for section_idx in range(num_sections):
        start_idx = section_idx * names_per_section
        end_idx = min(start_idx + names_per_section, len(names_for_normal_table))
        section_names = names_for_normal_table[start_idx:end_idx]
        datum_col = section_idx * 2 * (names_per_section + 1) + 1
        info_cell = ws.cell(row=1, column=datum_col)
        info_cell.value = f"Stundenliste - {calendar.month_name[month]} {year}"
        add_section(
            datum_col, 3, ws, year, month, section_names, person_lookup, db, master_db
        )
        next_column = datum_col + len(section_names) * 2 + 2

    for i, name in enumerate(names_for_extra_table):
        add_section(
            next_column + 2 + i * 6,
            3,
            ws,
            year,
            month,
            [name],
            person_lookup,
            db,
            master_db,
        )

    for col in range(1, next_column + 2 + len(names_for_extra_table) * 6):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return wb


def export_to_excel_top_to_bottom(
    year: int,
    month: int,
    db: Database,
    master_db: MasterDataDatabase,
    filename: str = None,
):
    if filename is None:
        filename = f"stundenliste_{year}_{month:02d}.xlsx"

    wb = build_workbook_top_to_bottom(year, month, db, master_db)
    if wb is None:
        return False

    try:
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False


def add_section(
    col,
    row,
    ws,
    year,
    month,
    section_names,
    person_lookup,
    db: Database,
    master_db: MasterDataDatabase,
):
    add_datum_header(col, 3, ws, year, month)
    num_days = calendar.monthrange(year, month)[1]
    name_to_col_map = {}

    for i, name in enumerate(section_names):
        name_col = col + 2 + i * 2
        name_to_col_map[name] = name_col

        # Write name (1x2 merged cell in row 3)
        ws.merge_cells(
            start_row=3, start_column=name_col, end_row=3, end_column=name_col + 1
        )

        name_cell = ws.cell(row=3, column=name_col)
        name_cell.value = name
        name_cell.alignment = Alignment(horizontal="center", vertical="center")
        name_cell.font = Font(bold=True)
        person_data = person_lookup.get(name, {})
        kein_verpflegung = bool(person_data.get("kein_verpflegungsgeld", 0))

        if not kein_verpflegung:
            name_cell.fill = openpyxl.styles.PatternFill(
                start_color=UNTER_8H_COLOR, end_color=UNTER_8H_COLOR, fill_type="solid"
            )

        # Apply thick border to name header
        set_create_border(
            min_row=3,
            max_row=4,
            min_col=name_col,
            max_col=name_col + 1,
            side_style=Side(style="thick"),
            ws=ws,
        )

        # Write "Std." and "Bst." in row 4
        std_cell = ws.cell(row=4, column=name_col)
        std_cell.value = "Std."
        std_cell.alignment = Alignment(horizontal="center", vertical="center")

        bst_cell = ws.cell(row=4, column=name_col + 1)
        bst_cell.value = "Bst."
        bst_cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 5
    for day in range(1, num_days + 1):
        arbeits_entries = {
            name: db.get_arbeitsstunden_for_day(year, month, day, name)
            for name in section_names
        }
        max_entries = max(
            max([len(entries) for entries in arbeits_entries.values()]), 1
        )
        for j in range(max_entries):
            date_cell = ws.cell(row=row + j, column=col)
            ws.merge_cells(
                start_row=row + j, start_column=col, end_row=row + j, end_column=col + 1
            )

            date_cell.value = f"{day}."
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
            # Color date cell if it's a weekend or holiday
            color_cell_weekend(
                col, row + j, ws, year, month, day, 2 * len(section_names) + 2
            )

        if is_holiday(year, month, day) and not is_weekend(year, month, day):
            for name in arbeits_entries:
                person_data = person_lookup.get(name, {})
                worker_type = person_data.get("worker_type", "Fest")
                std_cell_data = ws.cell(row=row + j, column=name_to_col_map[name])
                bst_cell_data = ws.cell(row=row + j, column=name_to_col_map[name] + 1)
                std_cell_data.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                bst_cell_data.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                if (
                    worker_type == WorkerTypes.Gewerblich
                    and not person_lookup[name]["keine_feiertagssstunden"]
                ):
                    std_cell_data.value = "F"
                    bst_cell_data.value = "940"
                elif (
                    worker_type == WorkerTypes.Fest
                    and sum(
                        e.get("stunden", 0)
                        for e in person_data.get("arbeits_entries", [])
                    )
                    > 0
                ):
                    weekly_hours = person_data.get("weekly_hours", 0.0)
                    std_cell_data.number_format = "0.00"
                    std_cell_data.value = weekly_hours / 5.0
                    bst_cell_data.value = "F"

        for name in arbeits_entries:
            meta_data = db.get_metadata_by_date(year, month, day, name)
            if meta_data is None:
                meta_data = {}
            person_data = person_lookup.get(name, {})
            worker_type = person_data.get("worker_type", "Fest")
            kein_verpflegung = bool(person_data.get("kein_verpflegungsgeld", 0))

            for j, entry in enumerate(arbeits_entries[name]):
                std_cell_data = ws.cell(row=row + j, column=name_to_col_map[name])
                bst_cell_data = ws.cell(row=row + j, column=name_to_col_map[name] + 1)
                std_cell_data.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                bst_cell_data.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                kostenstelle = entry.get("kostenstelle", "")
                if meta_data.get("kg_8h", False) and not kein_verpflegung:
                    std_cell_data.fill = openpyxl.styles.PatternFill(
                        start_color=UNTER_8H_COLOR,
                        end_color=UNTER_8H_COLOR,
                        fill_type="solid",
                    )
                    bst_cell_data.fill = openpyxl.styles.PatternFill(
                        start_color=UNTER_8H_COLOR,
                        end_color=UNTER_8H_COLOR,
                        fill_type="solid",
                    )
                skug_value = meta_data.get("skug")
                try:
                    skug_value = float(skug_value)
                except (TypeError, ValueError):
                    skug_value = 0.0
                if skug_value > 1:
                    std_cell_data.fill = openpyxl.styles.PatternFill(
                        start_color=SKUG_COLOR, end_color=SKUG_COLOR, fill_type="solid"
                    )
                    bst_cell_data.fill = openpyxl.styles.PatternFill(
                        start_color=SKUG_COLOR, end_color=SKUG_COLOR, fill_type="solid"
                    )
                if meta_data.get("travel_status", False):
                    std_cell_data.fill = openpyxl.styles.PatternFill(
                        start_color=AN_AB_COLOR,
                        end_color=AN_AB_COLOR,
                        fill_type="solid",
                    )
                    bst_cell_data.fill = openpyxl.styles.PatternFill(
                        start_color=AN_AB_COLOR,
                        end_color=AN_AB_COLOR,
                        fill_type="solid",
                    )

                if kostenstelle == "Krank":
                    std_cell_data.value = f"Krank"
                elif kostenstelle in ["940", "900"]:
                    std_cell_data.value = "F"
                    bst_cell_data.value = kostenstelle
                else:
                    std_cell_data.value = entry.get("stunden", 0)
                    std_cell_data.number_format = "0.00"
                    bst_cell_data.value = int(kostenstelle.split(" - ")[0])
        row += max_entries

    # Thick border around dates
    for i in range(len(section_names) + 1):
        set_create_border(
            min_row=5,
            max_row=row - 1,
            min_col=col + i * 2,
            max_col=col + i * 2 + 1,
            side_style=Side(style="thick"),
            ws=ws,
        )

    # Add summary rows under this section
    add_summary_rows(col, row, ws)
    fill_summary_rows(
        col + 2, row, ws, section_names, person_lookup, year, month, master_db, db
    )
    add_legend(col, row + len(summary_labels), ws)
    max_row = row + len(summary_labels) - 1
    max_col = col + len(section_names) * 2 + 1
    addLattice(3, max_row, col, max_col, ws)


def color_cell_weekend(col, row, ws, year, month, day, num_cols):
    if is_weekend(year, month, day) or is_holiday(year, month, day):
        for col_ in range(col, col + num_cols):
            ws.cell(row=row, column=col_).fill = openpyxl.styles.PatternFill(
                start_color=FREE_DAY_COLOR, end_color=FREE_DAY_COLOR, fill_type="solid"
            )


def add_datum_header(col, row, ws, year, month):
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + 1)
    datum_cell = ws.cell(row=row, column=col)
    datum_cell.value = "Datum"
    datum_cell.alignment = Alignment(horizontal="center", vertical="center")
    datum_cell.font = Font(bold=True)
    for row_ in range(row, row + 2):
        for col_ in range(col, col + 2):
            ws.cell(row=row_, column=col_).border = thick_border


def add_summary_rows(col, row, ws):
    summary_start_row = row
    # Apply thick border to summary labels
    set_create_border(
        min_row=summary_start_row,
        max_row=summary_start_row + len(summary_labels) - 1,
        min_col=col,
        max_col=col,
        side_style=Side(style="thick"),
        ws=ws,
    )

    # Write summary labels under Datum column
    for idx, label in enumerate(summary_labels):
        row = summary_start_row + idx
        label_cell = ws.cell(row=row, column=col)
        label_cell.value = label
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)


def fill_summary_rows(
    col, row, ws, section_names, person_lookup, year, month, master_db, db
):
    # Calculate and write summary values for each name
    for name_idx, name in enumerate(section_names):
        name_col = col + (name_idx * 2)

        # Apply thick border to summary numbers
        set_create_border(
            min_row=row,
            max_row=row + len(summary_labels) - 1,
            min_col=name_col,
            max_col=name_col + 1,
            side_style=Side(style="thick"),
            ws=ws,
        )
        person_data = person_lookup.get(name, {})
        worker_type = person_data.get("worker_type", "Fest")
        kein_verpflegung = bool(person_data.get("kein_verpflegungsgeld", 0))
        keine_feiertag = bool(person_data.get("keine_feiertagssstunden", 0))
        weekly_hours = person_data.get("weekly_hours", 0.0)

        # Get SKUG settings for calculating Feiertag hours
        # skug_settings = master_db.get_skug_settings()

        # Calculate totals
        if worker_type == WorkerTypes.Fest:
            urlaubsstunden = get_days_of_urlaub(name, month, year, db)
        else:
            urlaubsstunden = get_hours_of_urlaub(name, month, year, db)
        if worker_type == WorkerTypes.Fest:
            krankstunden = get_days_of_krank(name, month, year, db)
        else:
            krankstunden = get_hours_of_krank(name, month, year, db)

        if keine_feiertag:
            feiertag = 0
        elif worker_type == WorkerTypes.Fest:
            feiertag = get_days_of_feiertag(name, month, year)
        else:
            feiertag = get_hours_of_feiertag(
                name, month, year, master_db.get_skug_settings(), person_data
            )

        gesamtstunden = (
            sum(e.get("stunden", 0) for e in person_data.get("arbeits_entries", []))
            - get_hours_of_urlaub(name, month, year, db)
            - get_hours_of_krank(name, month, year, db)
        )
        skug_total = (
            get_skug_hours_for_name(name, month, year, db)
            if month in [12, 1, 2, 3]
            else 0
        )
        summe = (
            gesamtstunden
            + get_hours_of_feiertag(
                name, month, year, master_db.get_skug_settings(), person_data
            )
            + skug_total
            + get_hours_of_urlaub(name, month, year, db)
            + get_hours_of_krank(name, month, year, db)
        )
        if worker_type == WorkerTypes.Fest and gesamtstunden == 0:
            summe = 0
        mehr_minder = summe - get_normal_hours_per_month(year, month, master_db)
        if kein_verpflegung:
            v_zuschuss = 0
        else:
            v_zuschuss = get_verpflegungsgeld_for_name(name, month, year, master_db, db)
        summary_values = [
            gesamtstunden,
            feiertag,
            urlaubsstunden,
            krankstunden,
            skug_total,
            summe,
            mehr_minder,
            v_zuschuss,
        ]

        if worker_type == WorkerTypes.Gewerblich:
            create_zeitarbeiter_summary(
                ws, person_lookup, name, summary_values, row, name_col
            )
        elif worker_type == WorkerTypes.Fest:
            create_fest_summary(
                ws,
                name,
                month,
                year,
                summary_values,
                row,
                name_col,
                worker_type,
                master_db,
                db,
                weekly_hours,
            )


def add_legend(col, row, ws):
    cell = ws.cell(row=row, column=col)
    cell.value = "Wochenende/Feiertag"
    cell.font = Font(italic=True, color=FREE_DAY_COLOR)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws.cell(row=row + 1, column=col)
    cell.value = "diesen Tag mit SKUG auffüllen"
    cell.font = Font(italic=True, color=SKUG_COLOR)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws.cell(row=row + 2, column=col)
    cell.value = "weniger oder gleich als 8 Stunden von zu Hause abwesend"
    cell.font = Font(italic=True, color=UNTER_8H_COLOR)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws.cell(row=row + 3, column=col)
    cell.value = "An+Ab/>24"
    cell.font = Font(italic=True, color=AN_AB_COLOR)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def export_to_excel(
    year: int,
    month: int,
    db: Database,
    master_db: MasterDataDatabase,
    filename: str = None,
):
    """
    To be deleted soon
    """
    if filename is None:
        filename = f"stundenliste_{year}_{month:02d}.xlsx"

    # Get all entries for the month
    entries = db.get_entries_by_date(year, month)

    if not entries:
        print(f"No data to export for {year}-{month:02d}")
        return False

    # Get number of days in month
    num_days = calendar.monthrange(year, month)[1]

    # Get unique names from entries, sorted
    unique_names = (
        master_db.get_all_names_list()
    )  # sorted(set(entry['name'] for entry in entries))

    # Create a lookup for person data
    all_persons = master_db.get_all_names()
    person_lookup = {p["name"]: p for p in all_persons}

    if not unique_names:
        print("No names found in entries")
        return False

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    # Define thick border style
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

    # Track current column position
    current_col = 1  # Start at column A
    names_per_section = 9

    # Calculate how many sections we need
    num_sections = (len(unique_names) + names_per_section - 1) // names_per_section

    for section_idx in range(num_sections):
        # Get names for this section
        start_idx = section_idx * names_per_section
        end_idx = min(start_idx + names_per_section, len(unique_names))
        section_names = unique_names[start_idx:end_idx]

        # Write "Datum" header (2x2 merged cell starting at row 3)
        datum_col = current_col
        ws.merge_cells(
            start_row=3, start_column=datum_col, end_row=4, end_column=datum_col + 1
        )
        datum_cell = ws.cell(row=3, column=datum_col)
        datum_cell.value = "Datum"
        datum_cell.alignment = Alignment(horizontal="center", vertical="center")
        datum_cell.font = Font(bold=True)

        info_cell = ws.cell(row=1, column=datum_col)
        info_cell.value = f"Stundenliste - {calendar.month_name[month]} {year}"

        # Apply thick border to Datum header
        for row in range(3, 5):
            for col in range(datum_col, datum_col + 2):
                ws.cell(row=row, column=col).border = thick_border

        # Write dates (starting at row 5)
        for day in range(1, num_days + 1):
            row = 5 + day - 1
            ws.merge_cells(
                start_row=row,
                start_column=datum_col,
                end_row=row,
                end_column=datum_col + 1,
            )
            date_cell = ws.cell(row=row, column=datum_col)
            date_cell.value = f"{day}."
            date_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Color date cell if it's a weekend or holiday
            if is_weekend(year, month, day) or is_holiday(year, month, day):
                for col in range(datum_col, datum_col + 2):
                    ws.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                        start_color=FREE_DAY_COLOR,
                        end_color=FREE_DAY_COLOR,
                        fill_type="solid",
                    )

        # Apply thick border around all dates
        set_create_border(
            min_row=5,
            max_row=5 + num_days - 1,
            min_col=datum_col,
            max_col=datum_col + 1,
            side_style=Side(border_style="thick"),
            ws=ws,
        )

        # Move to name columns
        current_col += 2

        # Write each name column
        for name in section_names:
            name_col = current_col

            # Write name (1x2 merged cell in row 3)
            ws.merge_cells(
                start_row=3, start_column=name_col, end_row=3, end_column=name_col + 1
            )
            name_cell = ws.cell(row=3, column=name_col)
            name_cell.value = name
            name_cell.alignment = Alignment(horizontal="center", vertical="center")
            name_cell.font = Font(bold=True)
            person_data = person_lookup.get(name, {})
            worker_type = person_data.get("worker_type", "Fest")
            kein_verpflegung = bool(person_data.get("kein_verpflegungsgeld", 0))
            keine_feiertag = bool(person_data.get("keine_feiertagssstunden", 0))

            if not kein_verpflegung:
                name_cell.fill = openpyxl.styles.PatternFill(
                    start_color=UNTER_8H_COLOR,
                    end_color=UNTER_8H_COLOR,
                    fill_type="solid",
                )

            # Apply thick border to name header
            set_create_border(
                min_row=3,
                max_row=4,
                min_col=name_col,
                max_col=name_col + 1,
                side_style=Side(style="thick"),
                ws=ws,
            )

            # Write "Std." and "Bst." in row 4
            std_cell = ws.cell(row=4, column=name_col)
            std_cell.value = "Std."
            std_cell.alignment = Alignment(horizontal="center", vertical="center")

            bst_cell = ws.cell(row=4, column=name_col + 1)
            bst_cell.value = "Bst."
            bst_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply thick border around Std./Bst. Data
            set_create_border(
                min_row=5,
                max_row=5 + num_days - 1,
                min_col=name_col,
                max_col=name_col + 1,
                side_style=Side(style="thick"),
                ws=ws,
            )
            # Fill in data for each day
            for day in range(1, num_days + 1):
                row = 5 + day - 1

                # Check if this day is a weekend or holiday
                is_free_day = is_weekend(year, month, day) or is_holiday(
                    year, month, day
                )
                is_bank_holiday = is_holiday(year, month, day) and not is_weekend(
                    year, month, day
                )

                std_cell_data = ws.cell(row=row, column=name_col)
                bst_cell_data = ws.cell(row=row, column=name_col + 1)

            # Move to next name
            current_col += 2

        # Add summary rows under this section
        summary_start_row = 5 + num_days
        summary_labels = [
            "Gesamtstunden",
            "Feiertag",
            "Urlaubsstunden",
            "Krankstunden",
            "SKUG",
            "Summe",
            "Mehr-/Minderstd",
            "V.-Zuschuss [€]",
        ]

        # Apply thick border to summary labels
        set_create_border(
            min_row=summary_start_row,
            max_row=summary_start_row + len(summary_labels) - 1,
            min_col=datum_col,
            max_col=datum_col,
            side_style=Side(style="thick"),
            ws=ws,
        )

        # Write summary labels under Datum column
        for idx, label in enumerate(summary_labels):
            row = summary_start_row + idx
            label_cell = ws.cell(row=row, column=datum_col)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            # Merge across both Datum columns
            ws.merge_cells(
                start_row=row,
                start_column=datum_col,
                end_row=row,
                end_column=datum_col + 1,
            )

        # Add legend for colors
        row = summary_start_row + len(summary_labels)

        cell = ws.cell(row=row, column=datum_col)
        cell.value = "Wochenende/Feiertag"
        cell.font = Font(italic=True, color=FREE_DAY_COLOR)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        cell = ws.cell(row=row + 1, column=datum_col)
        cell.value = "diesen Tag mit SKUG auffüllen"
        cell.font = Font(italic=True, color=SKUG_COLOR)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        cell = ws.cell(row=row + 1, column=datum_col + 5)
        cell.value = "weniger oder gleich als 8 Stunden von zu Hause abwesend"
        cell.font = Font(italic=True, color=UNTER_8H_COLOR)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        cell = ws.cell(row=row + 1, column=datum_col + 11)
        cell.value = "An+Ab/>24"
        cell.font = Font(italic=True, color=AN_AB_COLOR)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Calculate and write summary values for each name
        for name_idx, name in enumerate(section_names):
            name_col = datum_col + 2 + (name_idx * 2)

            # Apply thick border to summary numbers
            set_create_border(
                min_row=summary_start_row,
                max_row=summary_start_row + len(summary_labels) - 1,
                min_col=name_col,
                max_col=name_col + 1,
                side_style=Side(style="thick"),
                ws=ws,
            )

            # Get all entries for this name
            name_entries = [e for e in entries if e["name"] == name]

            person_data = person_lookup.get(name, {})
            worker_type = person_data.get("worker_type", "Fest")
            kein_verpflegung = bool(person_data.get("kein_verpflegungsgeld", 0))
            keine_feiertag = bool(person_data.get("keine_feiertagssstunden", 0))
            weekly_hours = person_data.get("weekly_hours", 0.0)

            # Get SKUG settings for calculating Feiertag hours
            skug_settings = master_db.get_skug_settings()

            # Calculate totals
            gesamtstunden = 0
            urlaubsstunden = 0
            krankstunden = 0
            skug_total = 0
            feiertag = 0
            summe = 0
            mehr_minder = 0
            v_zuschuss = 0

            summary_values = [
                gesamtstunden,
                feiertag,
                urlaubsstunden,
                krankstunden,
                skug_total,
                summe,
                mehr_minder,
                v_zuschuss,
            ]

            if worker_type == WorkerTypes.Gewerblich:
                create_zeitarbeiter_summary(
                    ws, person_lookup, name, summary_values, summary_start_row, name_col
                )
            elif worker_type == WorkerTypes.Fest:
                create_fest_summary(
                    ws,
                    name,
                    month,
                    year,
                    summary_values,
                    summary_start_row,
                    name_col,
                    worker_type,
                    master_db,
                    db,
                    weekly_hours,
                )
        current_col += 2

    for col in range(1, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 12

    max_row = 4 + num_days + len(summary_labels)
    max_col = len(unique_names) * 2 + num_sections * 2
    addLattice(3, max_row, 1, max_col, ws)
    try:
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False


def create_zeitarbeiter_summary(
    ws: Workbook,
    person_lookup,
    name,
    summary_values: list,
    summary_start_row: int = None,
    name_col: int = None,
):
    for idx, value in enumerate(summary_values):
        row = summary_start_row + idx
        value_cell = ws.cell(row=row, column=name_col)
        value_cell.value = value if value != 0 else ""
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = "0.00"

        if idx == 6:
            person_data = person_lookup.get(name, {})
            keine_feiertag = bool(person_data.get("kein_fzk", 0))
            if keine_feiertag:
                ws.merge_cells(
                    start_row=row,
                    start_column=name_col,
                    end_row=row,
                    end_column=name_col + 1,
                )
                value_cell = ws.cell(row=row, column=name_col)
                value_cell.value = "Kein FZK"
                value_cell.alignment = Alignment(horizontal="center", vertical="center")
                set_create_border(
                    min_row=row,
                    max_row=row,
                    min_col=name_col,
                    max_col=name_col + 1,
                    side_style=Side(style="thick"),
                    ws=ws,
                )


def create_fest_summary(
    ws: Workbook,
    name,
    month,
    year,
    summary_values: list,
    summary_start_row: int,
    name_col: int,
    worker_type: WorkerTypes,
    master_db: MasterDataDatabase,
    db: Database,
    weekly_hours: float = 0.0,
):
    stunden = summary_values[0]
    for idx, value in enumerate(summary_values):
        row = summary_start_row + idx

        # Standard value writing first (can be overwritten)
        value_cell = ws.cell(row=row, column=name_col)
        value_cell.value = value if value != 0 else ""
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if idx == 0:
            value_cell.number_format = "0.00"
        if idx == 1:  # Feiertag
            value_cell = ws.cell(row=row, column=name_col + 1)
            value_cell.value = "Tage"
        if idx == 2:  # Urlaubsstunden
            value_cell = ws.cell(row=row, column=name_col + 1)
            value_cell.value = "Tage"

            value_cell = ws.cell(row=row, column=name_col)
            value_cell.value = get_days_of_urlaub(name, month, year, db)
        if idx == 3:  # Krankstunden
            value_cell = ws.cell(row=row, column=name_col + 1)
            value_cell.value = "Tage"

            value_cell = ws.cell(row=row, column=name_col)
            value_cell.value = get_days_of_krank(name, month, year, db)
        if idx == 5:
            value_cell.number_format = "0.00"
        if idx == 6:  # Mehr-/Minderstd
            # If weekly_hours > 0, we show the calculated value (already in summary_values[6])
            # If NOT weekly_hours > 0, we do the old merge thing
            if stunden <= 0:
                ws.merge_cells(
                    start_row=row,
                    start_column=name_col,
                    end_row=row,
                    end_column=name_col + 1,
                )
                value_cell = ws.cell(row=row, column=name_col)
                value_cell.value = worker_type
                value_cell.alignment = Alignment(horizontal="center", vertical="center")
                set_create_border(
                    min_row=row,
                    max_row=row,
                    min_col=name_col,
                    max_col=name_col + 1,
                    side_style=Side(style="thick"),
                    ws=ws,
                )
            else:
                # Just ensure formatting is correct for the value
                value_cell = ws.cell(row=row, column=name_col)
                value_cell.value = value
                value_cell.number_format = "0.00"
                value_cell.alignment = Alignment(horizontal="center", vertical="center")
