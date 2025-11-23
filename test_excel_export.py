import os
import sqlite3
import datetime
from database import Database
from master_data import MasterDataDatabase
from excel_export import export_to_excel

DB_FILE = "test_db.db"
MASTER_DB_FILE = "test_master_data.db"
EXPORT_FILE = "test_export.xlsx"

def setup_dbs():
    if os.path.exists(DB_FILE): os.remove(DB_FILE)
    if os.path.exists(MASTER_DB_FILE): os.remove(MASTER_DB_FILE)
    
    db = Database(DB_FILE)
    master_db = MasterDataDatabase(MASTER_DB_FILE)
    
    # Add test persons
    # 1. Normal Fest (should have Verpflegung, Feiertag SKUG, Normal Summe)
    master_db.add_name("Normal Fest", "Fest")
    
    # 2. Fest with Weekly Hours (should have Feiertag=Weekly/5, Fixed Summe, Mehr/Minder calc)
    master_db.add_name("Weekly Fest", "Fest", weekly_hours=40.0)
    
    # 3. No Verpflegung (should have 0 Verpflegung)
    master_db.add_name("No Verpflegung", "Fest", kein_verpflegungsgeld=True)
    
    # 4. No Feiertag (should have 0 Feiertag)
    master_db.add_name("No Feiertag", "Fest", keine_feiertagssstunden=True)
    
    return db, master_db

def test_export():
    db, master_db = setup_dbs()
    
    # Add entries for May 2024 (has holidays: 1.5, 9.5, 20.5, 30.5)
    # 1.5.2024 is Wednesday (Feiertag)
    # 9.5.2024 is Thursday (Feiertag)
    
    year = 2024
    month = 5
    
    # Add some hours
    # Helper to add entry
    def add_entry(name, day, hours, baustelle):
        data = {
            'Jahr': year,
            'Monat': month,
            'Tag': day,
            'Name': name,
            'Wochentag': 'Thu', # 2.5.2024 is Thursday
            'Stunden': hours,
            'Urlaub': '',
            'Krank': '',
            'unter_8h': False,
            'SKUG': '',
            'Baustelle': baustelle
        }
        db.add_or_update_entry(data)

    # Normal Fest: 8h on 2.5 (Thu)
    add_entry("Normal Fest", 2, "8", "123 - Baustelle A")
    
    # Weekly Fest: 8h on 2.5 (Thu)
    add_entry("Weekly Fest", 2, "8", "123 - Baustelle A")
    
    # No Verpflegung: 8h on 2.5 (Thu)
    add_entry("No Verpflegung", 2, "8", "123 - Baustelle A")
    
    # No Feiertag: 8h on 2.5 (Thu)
    add_entry("No Feiertag", 2, "8", "123 - Baustelle A")
    
    # Add Baustelle with Verpflegungsgeld
    master_db.add_baustelle("123", "Baustelle A", verpflegungsgeld=10.0)
    
    print("Exporting to Excel...")
    success = export_to_excel(year, month, db, master_db, EXPORT_FILE)
    assert success
    
    print("Export successful. Please verify manually or I can add openpyxl checks here.")
    
    # Verify with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(EXPORT_FILE)
    ws = wb.active
    
    # Find columns for each person
    # Layout: Datum (A,B), Name1 (C,D), Name2 (E,F), ...
    # Names are sorted alphabetically: 
    # No Feiertag, No Verpflegung, Normal Fest, Weekly Fest
    
    names = master_db.get_all_names_list()
    print(f"Names order: {names}")
    
    # Helper to find col index for name
    def get_col_for_name(name):
        idx = names.index(name)
        return 3 + (idx * 2)
    
    # Verify No Verpflegung
    col = get_col_for_name("No Verpflegung")
    # Summary rows start after days (31 days + 5 header rows = 36)
    # Summary labels start at row 36
    # V.-Zuschuss is the last one (index 8) -> row 36 + 8 = 44
    v_zuschuss_cell = ws.cell(row=44, column=col)
    print(f"No Verpflegung V-Zuschuss: {v_zuschuss_cell.value}")
    assert v_zuschuss_cell.value in [0, 0.0, "", None]
    
    # Verify Normal Fest Verpflegung (should be > 0)
    col = get_col_for_name("Normal Fest")
    v_zuschuss_cell = ws.cell(row=44, column=col)
    print(f"Normal Fest V-Zuschuss: {v_zuschuss_cell.value}")
    assert v_zuschuss_cell.value == 10.0 # 1 day * 10.0
    
    # Verify No Feiertag
    col = get_col_for_name("No Feiertag")
    # Feiertag is index 1 -> row 37
    feiertag_cell = ws.cell(row=37, column=col)
    print(f"No Feiertag Hours: {feiertag_cell.value}")
    assert feiertag_cell.value in [0, 0.0, "", None]
    
    # Verify Weekly Fest Feiertag
    # May 2024 in SH has 3 holidays (1.5, 9.5, 20.5). 30.5 (Fronleichnam) is NOT a holiday in SH.
    # Weekly hours = 40.0 -> 8.0 per day.
    # Should be 3 * 8.0 = 24.0
    col = get_col_for_name("Weekly Fest")
    feiertag_cell = ws.cell(row=37, column=col)
    print(f"Weekly Fest Feiertag Hours: {feiertag_cell.value}")
    assert float(feiertag_cell.value) == 24.0
    
    # Verify Weekly Fest Summe
    # Summe = 40 * 52 / 12 = 173.333...
    col = get_col_for_name("Weekly Fest")
    # Summe is index 5 -> row 41
    summe_cell = ws.cell(row=41, column=col)
    print(f"Weekly Fest Summe: {summe_cell.value}")
    assert abs(float(summe_cell.value) - 173.333) < 0.01
    
    # Verify Weekly Fest Mehr/Minder
    # Actual = 8 (worked) + 24 (feiertag) = 32
    # Target = 173.333
    # Mehr/Minder = 32 - 173.333 = -141.333
    col = get_col_for_name("Weekly Fest")
    # Mehr/Minder is index 6 -> row 42
    mehr_minder_cell = ws.cell(row=42, column=col)
    print(f"Weekly Fest Mehr/Minder: {mehr_minder_cell.value}")
    assert abs(float(mehr_minder_cell.value) - (32 - 173.333)) < 0.01
    
    print("Verification passed!")
    
    # Cleanup
    if os.path.exists(DB_FILE): os.remove(DB_FILE)
    if os.path.exists(MASTER_DB_FILE): os.remove(MASTER_DB_FILE)
    if os.path.exists(EXPORT_FILE): os.remove(EXPORT_FILE)

if __name__ == "__main__":
    test_export()
