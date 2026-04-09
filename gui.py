import tkinter as tk
from tkinter import ttk, messagebox
from concurrent.futures import ThreadPoolExecutor
from tksheet import Sheet
from database import Database
from excel_export import (
    export_to_excel,
    export_to_excel_top_to_bottom,
    build_workbook_top_to_bottom,
)
from openpyxl.utils import get_column_letter
from utils import validate_required_fields, get_next_day_skip_weekend, get_next_day
from utils import get_weekday_abbr, parse_date_range, parse_multiple_names
from utils import validate_days_in_month, calculate_skug, get_effective_fahrzeit
from utils import (
    handle_krank_urlaub,
    try_load_existing_entry,
    check_arbeitsstunden,
    determine_kg_8h_flag,
)
from datetime import datetime, timedelta
from master_data import MasterDataDatabase
from manager_dialogs import NameManagerDialog, BaustelleManagerDialog
from autocomplete import AutocompleteEntry, BaustelleAutocomplete
from settings_dialog import Settings, SettingsDialog
from datatypes import TravelStatus, WorkerTypes
from entry_service import EntryService


class ExcelPreviewWindow:
    def __init__(
        self,
        parent,
        on_close,
        on_edit=None,
        on_apply=None,
        on_flag=None,
        on_reset=None,
        on_modified=None,
    ):
        self.parent = parent
        self.on_close = on_close
        self.on_edit = on_edit
        self.on_apply = on_apply
        self.on_flag = on_flag
        self.on_reset = on_reset
        self.on_modified = on_modified
        self.cell_map = {}
        self.preview_year = None
        self.preview_month = None
        self._suppress_edit_events = False
        self.original_values = {}
        self.base_status_text = ""
        self.window = tk.Toplevel(parent)
        self.window.title("Excel Vorschau")
        self.window.geometry("1200x700")
        self.window.minsize(800, 400)
        self.window.resizable(True, True)
        self.window.protocol("WM_DELETE_WINDOW", self.close)

        top_frame = tk.Frame(self.window)
        top_frame.pack(fill=tk.X, padx=10, pady=(10, 0))

        self.apply_button = tk.Button(
            top_frame, text="Anwenden", command=self._on_apply_click
        )
        self.apply_button.pack(side=tk.LEFT, padx=(0, 8))

        self.reset_button = tk.Button(
            top_frame, text="Reset", command=self._on_reset_click
        )
        self.reset_button.pack(side=tk.LEFT, padx=(0, 8))

        self.title_label = tk.Label(top_frame, text="", anchor="w")
        self.title_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.status_label = tk.Label(self.window, text="", anchor="w", fg="gray")
        self.status_label.pack(fill=tk.X, padx=10, pady=(4, 8))

        tree_frame = tk.Frame(self.window)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.sheet = Sheet(tree_frame, data=[[]])
        self.sheet.pack(fill=tk.BOTH, expand=True)
        self._safe_sheet_call("enable_bindings", "all")
        self._bind_edit_events()

    def close(self):
        if callable(self.on_close):
            self.on_close()
        self.window.destroy()

    def is_open(self):
        return self.window.winfo_exists()

    def clear(self):
        self._set_sheet_data([], ["#"])

    def show_message(self, title_text, status_text=""):
        self.title_label.config(text=title_text)
        self.status_label.config(text=status_text)
        self.base_status_text = status_text
        self.clear()
        self._safe_sheet_call("headers", ["#"])

    def render_workbook(self, workbook, year, month, cell_map=None):
        self._suppress_edit_events = True
        if workbook is None:
            self.show_message("Excel Vorschau", "Keine Daten zum Anzeigen vorhanden.")
            self._suppress_edit_events = False
            return

        ws = workbook.active
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            self.show_message(
                f"Excel Vorschau {month:02d}/{year}",
                "Leere Tabelle.",
            )
            self._suppress_edit_events = False
            return

        headers = ["#"] + [get_column_letter(i) for i in range(1, max_col + 1)]
        data = []
        for row_idx in range(1, max_row + 1):
            row_values = [row_idx]
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_values.append(self._format_preview_value(cell))
            data.append(row_values)

        self._clear_all_highlights()
        self._set_sheet_data(data, headers)
        self._apply_dimensions(ws, data_col_offset=1)
        self._apply_merges(ws, data_col_offset=1)
        self._apply_styles(ws, data_col_offset=1)

        self.title_label.config(text=f"Excel Vorschau {month:02d}/{year}")
        self.base_status_text = (
            f"Blatt: {ws.title} | Zeilen: {max_row} | Spalten: {max_col}"
        )
        self.status_label.config(text=self.base_status_text)
        self.cell_map = cell_map or {}
        self.preview_year = year
        self.preview_month = month
        self.original_values = {}
        for r_idx, row_data in enumerate(data):
            for c_idx, cell_value in enumerate(row_data):
                self.original_values[(r_idx, c_idx)] = cell_value
        self._suppress_edit_events = False

    def _format_preview_value(self, cell):
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, (int, float)):
            number_format = str(getattr(cell, "number_format", "") or "").strip()
            if number_format == "0.00" or isinstance(value, float):
                return f"{float(value):.2f}"
        return str(value)

    def _clear_all_highlights(self):
        if hasattr(self.sheet, "dehighlight_all"):
            try:
                self.sheet.dehighlight_all()
                return
            except Exception:
                pass
        if hasattr(self.sheet, "dehighlight_cells"):
            for row, col in list(self.original_values.keys()):
                try:
                    self.sheet.dehighlight_cells(row, col)
                except Exception:
                    pass

    def set_pending_count(self, count):
        if not self.base_status_text:
            return
        if count:
            self.status_label.config(
                text=f"{self.base_status_text} | Änderungen: {count}"
            )
        else:
            self.status_label.config(text=self.base_status_text)

    def _bind_edit_events(self):
        try:
            self.sheet.extra_bindings([("edit_cell", self._on_sheet_edit)])
        except Exception:
            pass
        try:
            self.sheet.extra_bindings([("paste", self._on_sheet_paste)])
        except Exception:
            pass
        try:
            self.sheet.bind("<<SheetModified>>", self._on_sheet_modified)
        except Exception:
            pass
        self._bind_right_click_menu()

    def _on_apply_click(self):
        if callable(self.on_apply):
            self.on_apply()

    def _on_reset_click(self):
        if not messagebox.askyesno(
            "Änderungen verwerfen",
            "Nicht bestätigte Änderungen wirklich verwerfen?",
        ):
            return
        if hasattr(self, "on_reset") and callable(self.on_reset):
            self.on_reset()

    def _on_sheet_edit(self, event):
        if self._suppress_edit_events:
            return
        row, col, value = self._parse_edit_event(event)
        if row is None or col is None:
            return
        if callable(self.on_edit):
            self.on_edit(row, col, value)

    def _on_sheet_modified(self, event):
        if self._suppress_edit_events:
            return
        if hasattr(self.sheet, "get_last_event_data"):
            data = self.sheet.get_last_event_data()
        elif hasattr(self.sheet, "get_last_event"):
            data = self.sheet.get_last_event()
        else:
            return
        row, col, value = self._parse_edit_event(data)
        if row is None or col is None:
            if callable(self.on_modified):
                self.on_modified()
            return
        if callable(self.on_edit):
            self.on_edit(row, col, value)

    def _on_sheet_paste(self, event):
        if self._suppress_edit_events:
            return
        if callable(self.on_modified):
            self.on_modified()

    def _parse_edit_event(self, event):
        if isinstance(event, dict):
            row = event.get("row")
            col = event.get("column")
            value = event.get("text")
            if value is None:
                value = event.get("value")
            return row, col, value
        if isinstance(event, (list, tuple)):
            if len(event) >= 3:
                return event[0], event[1], event[2]
        return None, None, None

    def _on_right_click(self, event):
        row, col = self._get_row_col_from_event(event)
        if row is None or col is None:
            row, col = self._get_row_col_from_selection()
            if row is None or col is None:
                return
        if col <= 0:
            return

        wb_row = row + 1
        wb_col = col
        cell_info = self.cell_map.get((wb_row, wb_col))
        if not cell_info:
            return
        year = cell_info.get("year")
        month = cell_info.get("month")
        day = cell_info.get("day")
        name = cell_info.get("name")
        if not all([year, month, day, name]):
            return

        has_work_entry = self._day_has_preview_work_entry(year, month, day, name)

        menu = tk.Menu(self.window, tearoff=0)
        menu.add_command(
            label="Krank",
            command=lambda: self._emit_flag(
                year, month, day, name, "krank", True, row=row, col=col
            ),
        )
        menu.add_command(
            label="Urlaub",
            command=lambda: self._emit_flag(
                year, month, day, name, "urlaub", True, row=row, col=col
            ),
        )
        menu.add_separator()
        menu.add_command(
            label="Reise: Anreise",
            state="normal" if has_work_entry else "disabled",
            command=lambda: self._emit_flag(
                year, month, day, name, "travel_status", "Anreise", row=row, col=col
            ),
        )
        menu.add_command(
            label="Reise: Abreise",
            state="normal" if has_work_entry else "disabled",
            command=lambda: self._emit_flag(
                year, month, day, name, "travel_status", "Abreise", row=row, col=col
            ),
        )
        menu.add_command(
            label="Reise: 24h_away",
            state="normal" if has_work_entry else "disabled",
            command=lambda: self._emit_flag(
                year, month, day, name, "travel_status", "24h_away", row=row, col=col
            ),
        )
        menu.add_command(
            label="Reise: Entfernen",
            command=lambda: self._emit_flag(
                year, month, day, name, "travel_status", None, row=row, col=col
            ),
        )
        menu.add_separator()
        menu.add_command(
            label="Frühstück an",
            state="normal" if has_work_entry else "disabled",
            command=lambda: self._emit_flag(
                year, month, day, name, "fruehstueck", True, row=row, col=col
            ),
        )
        menu.add_command(
            label="Frühstück aus",
            command=lambda: self._emit_flag(
                year, month, day, name, "fruehstueck", False, row=row, col=col
            ),
        )
        menu.add_command(
            label="Mittag an",
            state="normal" if has_work_entry else "disabled",
            command=lambda: self._emit_flag(
                year, month, day, name, "mittag", True, row=row, col=col
            ),
        )
        menu.add_command(
            label="Mittag aus",
            command=lambda: self._emit_flag(
                year, month, day, name, "mittag", False, row=row, col=col
            ),
        )
        menu.add_separator()
        menu.add_command(
            label="No SKUG an",
            command=lambda: self._emit_flag(
                year, month, day, name, "no_skug", True, row=row, col=col
            ),
        )
        menu.add_command(
            label="No SKUG aus",
            command=lambda: self._emit_flag(
                year, month, day, name, "no_skug", False, row=row, col=col
            ),
        )
        menu.tk_popup(event.x_root, event.y_root)
        menu.grab_release()
        return "break"

    def _emit_flag(self, year, month, day, name, flag, value, row=None, col=None):
        if callable(self.on_flag):
            self.on_flag(year, month, day, name, flag, value, row, col)

    def _day_has_preview_work_entry(self, year, month, day, name):
        row_values = {}
        for (wb_row, wb_col), cell_info in self.cell_map.items():
            if (
                cell_info.get("year") == year
                and cell_info.get("month") == month
                and cell_info.get("day") == day
                and cell_info.get("name") == name
            ):
                if cell_info.get("entry_id") is not None:
                    return True
                sheet_row = wb_row - 1
                row_values.setdefault(sheet_row, {})[cell_info.get("field")] = (
                    self.get_cell_text(sheet_row, wb_col) or ""
                )
        for values in row_values.values():
            if (
                str(values.get("Stunden", "")).strip()
                and str(values.get("Kostenstelle", "")).strip()
            ):
                return True
        return False

    def _get_row_col_from_event(self, event):
        if hasattr(self.sheet, "get_row_col_from_event"):
            try:
                return self.sheet.get_row_col_from_event(event)
            except Exception:
                pass
        if hasattr(self.sheet, "get_row_at_event") and hasattr(
            self.sheet, "get_column_at_event"
        ):
            try:
                row = self.sheet.get_row_at_event(event)
                col = self.sheet.get_column_at_event(event)
                return row, col
            except Exception:
                pass
        return None, None

    def _get_row_col_from_selection(self):
        if hasattr(self.sheet, "get_selected_cells"):
            try:
                selected = self.sheet.get_selected_cells()
                if selected:
                    row, col = selected[0]
                    return row, col
            except Exception:
                pass
        if hasattr(self.sheet, "get_currently_selected"):
            try:
                current = self.sheet.get_currently_selected()
                if current and len(current) >= 2:
                    return current[0], current[1]
            except Exception:
                pass
        return None, None

    def _bind_right_click_menu(self):
        bindings = ["<Button-3>", "<ButtonRelease-3>"]
        if self.window.tk.call("tk", "windowingsystem") == "aqua":
            bindings.extend(["<Button-2>", "<Control-Button-1>"])
        targets = [self.sheet]
        if hasattr(self.sheet, "MT"):
            targets.append(self.sheet.MT)
        for target in targets:
            for sequence in bindings:
                try:
                    target.bind(sequence, self._on_right_click, add="")
                except Exception:
                    pass

    def get_cell_text(self, row, col):
        if hasattr(self.sheet, "get_cell_data"):
            return self.sheet.get_cell_data(row, col)
        if hasattr(self.sheet, "get_cell_value"):
            return self.sheet.get_cell_value(row, col)
        return ""

    def set_cell_text(self, row, col, value):
        if hasattr(self.sheet, "set_cell_data"):
            try:
                self.sheet.set_cell_data(row, col, value)
                return True
            except Exception:
                pass
        if hasattr(self.sheet, "set_cell_value"):
            try:
                self.sheet.set_cell_value(row, col, value)
                return True
            except Exception:
                pass
        return False

    def mark_changed_cell(self, row, col, is_changed):
        if self._suppress_edit_events:
            return
        if is_changed:
            self._safe_sheet_call(
                "highlight_cells",
                row=row,
                column=col,
                bg="#fff2cc",
            )
        else:
            if hasattr(self.sheet, "dehighlight_cells"):
                try:
                    self.sheet.dehighlight_cells(row, col)
                except Exception:
                    pass

    def clear_pending_highlights(self):
        if hasattr(self.sheet, "dehighlight_all"):
            try:
                self.sheet.dehighlight_all()
                return
            except Exception:
                pass
        if hasattr(self.sheet, "dehighlight_cells"):
            for row, col in list(self.original_values.keys()):
                try:
                    self.sheet.dehighlight_cells(row, col)
                except Exception:
                    pass

    def _set_sheet_data(self, data, headers):
        if not self._safe_sheet_call(
            "set_sheet_data",
            data,
            reset_col_positions=True,
            reset_row_positions=True,
        ):
            self._safe_sheet_call("set_sheet_data", data)

        if not self._safe_sheet_call("headers", headers):
            self._safe_sheet_call("set_header_data", headers)

    def _safe_sheet_call(self, method_name, *args, **kwargs):
        if not hasattr(self.sheet, method_name):
            return False
        try:
            getattr(self.sheet, method_name)(*args, **kwargs)
            return True
        except Exception:
            return False

    def _excel_color_to_hex(self, color):
        if color is None:
            return None
        rgb = getattr(color, "rgb", None)
        if rgb is None:
            return None
        if not isinstance(rgb, str):
            rgb = getattr(rgb, "value", None) or str(rgb)
        rgb = rgb.strip()
        if rgb.startswith("#"):
            rgb = rgb[1:]
        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) != 6:
            return None
        return f"#{rgb}"

    def _apply_dimensions(self, ws, data_col_offset):
        col_widths = []
        for col_idx in range(1, ws.max_column + 1):
            dim = ws.column_dimensions.get(get_column_letter(col_idx))
            width = getattr(dim, "width", None) if dim else None
            if width is None:
                col_widths.append(None)
            else:
                col_widths.append(int(width * 7 + 5))

        if any(w is not None for w in col_widths):
            self._safe_sheet_call(
                "set_column_widths",
                [60] + [w or 90 for w in col_widths],
            )

        row_heights = []
        for row_idx in range(1, ws.max_row + 1):
            dim = ws.row_dimensions.get(row_idx)
            height = getattr(dim, "height", None) if dim else None
            if height is None:
                row_heights.append(None)
            else:
                row_heights.append(int(height * 96 / 72))

        if any(h is not None for h in row_heights):
            self._safe_sheet_call(
                "set_row_heights",
                [h or 20 for h in row_heights],
            )

    def _apply_merges(self, ws, data_col_offset):
        for merged in ws.merged_cells.ranges:
            r0 = merged.min_row - 1
            c0 = merged.min_col - 1 + data_col_offset
            r1 = merged.max_row - 1
            c1 = merged.max_col - 1 + data_col_offset
            if hasattr(self.sheet, "span"):
                try:
                    self.sheet.span(r0, c0, r1, c1)
                    continue
                except TypeError:
                    try:
                        self.sheet.span(
                            r0,
                            c0,
                            r1 - r0 + 1,
                            c1 - c0 + 1,
                        )
                        continue
                    except Exception:
                        pass
                except Exception:
                    pass
            if hasattr(self.sheet, "span_cells"):
                try:
                    self.sheet.span_cells(r0, c0, r1, c1)
                except Exception:
                    pass

    def _apply_styles(self, ws, data_col_offset):
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                sheet_row = row_idx - 1
                sheet_col = col_idx - 1 + data_col_offset

                fill = getattr(cell, "fill", None)
                if fill and getattr(fill, "fill_type", None) == "solid":
                    fill_hex = self._excel_color_to_hex(
                        getattr(fill, "start_color", None)
                    )
                    if fill_hex:
                        self._safe_sheet_call(
                            "highlight_cells",
                            row=sheet_row,
                            column=sheet_col,
                            bg=fill_hex,
                        )

                font = getattr(cell, "font", None)
                if font is not None:
                    font_color = self._excel_color_to_hex(getattr(font, "color", None))
                    if font_color:
                        self._safe_sheet_call(
                            "highlight_cells",
                            row=sheet_row,
                            column=sheet_col,
                            fg=font_color,
                        )
                    if getattr(font, "bold", False) or getattr(font, "italic", False):
                        self._safe_sheet_call(
                            "set_cell_font",
                            sheet_row,
                            sheet_col,
                            bold=getattr(font, "bold", False),
                            italic=getattr(font, "italic", False),
                        )

                alignment = getattr(cell, "alignment", None)
                if alignment is not None:
                    horizontal = getattr(alignment, "horizontal", None)
                    if horizontal in {"center", "left", "right"}:
                        align_value = (
                            "center"
                            if horizontal == "center"
                            else "w"
                            if horizontal == "left"
                            else "e"
                        )
                        if not self._safe_sheet_call(
                            "set_cell_align",
                            sheet_row,
                            sheet_col,
                            align_value,
                        ):
                            self._safe_sheet_call(
                                "align_cells",
                                sheet_row,
                                sheet_col,
                                align=align_value,
                            )


class StundenEingabeGUI:
    def __init__(self, root):
        self.root = root
        self.db = Database()
        self.master_db = MasterDataDatabase()
        self.settings = Settings()
        self.entry_service = EntryService(self.db, self.master_db)
        self.edit_mode_active = False
        self.edit_entry_id = None
        self.edit_original_year = None
        self.edit_original_month = None
        self.edit_original_day = None
        self.edit_original_name = None
        self.preview_window = None
        self.preview_refresh_job = None
        self.preview_executor = ThreadPoolExecutor(
            max_workers=1, thread_name_prefix="excel_preview"
        )
        self.preview_task_future = None
        self.preview_pending_request = None
        self.preview_inflight_request_id = 0
        self.preview_request_seq = 0
        self.preview_pending_edits = {}
        self.preview_pending_flags = {}
        self.setup_window()
        self.create_widgets()
        self.setup_bindings()

    def setup_window(self):
        self.root.title("Stunden-Eingabe")
        self.root.geometry("1000x600")
        self.root.protocol("WM_DELETE_WINDOW", self.on_app_close)

    def create_widgets(self):
        paned_window = tk.PanedWindow(
            self.root,
            orient=tk.HORIZONTAL,
            sashwidth=5,
            sashrelief=tk.RAISED,
            bg="gray",
        )
        paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        input_frame = tk.Frame(paned_window)
        paned_window.add(input_frame, minsize=300)
        display_frame = tk.Frame(paned_window)
        paned_window.add(display_frame, minsize=400)
        self.create_input_fields(input_frame)
        self.create_data_displays(display_frame)

    def day_has_work_entry(self, year, month, day, name):
        return self.entry_service.day_has_work_entry(year, month, day, name)

    def create_input_fields(self, parent):
        tk.Label(parent, text="Jahr:").grid(row=0, column=0, sticky="e", padx=5, pady=2)
        self.entry_year = tk.Entry(parent)
        self.entry_year.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        tk.Label(parent, text="Monat:").grid(
            row=1, column=0, sticky="e", padx=5, pady=2
        )
        self.entry_month = tk.Entry(parent)
        self.entry_month.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        self.label_day = tk.Label(parent, text="Tag(e):*")
        self.label_day.grid(row=2, column=0, sticky="e", padx=5, pady=2)
        self.entry_day = tk.Entry(parent)
        self.entry_day.grid(row=2, column=1, padx=5, pady=2, sticky="ew")
        tk.Label(parent, text="(z.B. 3-7,9,11-13)", font=("Arial", 7), fg="gray").grid(
            row=2, column=2, sticky="w", padx=2
        )
        tk.Label(parent, text="Name(n):").grid(
            row=3, column=0, sticky="e", padx=5, pady=2
        )
        name_frame = tk.Frame(parent)
        name_frame.grid(row=3, column=1, padx=5, pady=2, sticky="ew")
        self.entry_name = tk.Entry(name_frame)
        self.entry_name.pack(side=tk.LEFT, fill=tk.X, expand=True)
        btn_name_manager = tk.Button(
            name_frame, text="⚙", width=2, command=self.open_name_manager
        )
        btn_name_manager.pack(side=tk.LEFT, padx=(2, 0))
        tk.Label(parent, text="(z.B. Max, Anna)", font=("Arial", 7), fg="gray").grid(
            row=3, column=2, sticky="w", padx=2
        )
        tk.Label(parent, text="Stunden:").grid(
            row=4, column=0, sticky="e", padx=5, pady=2
        )
        self.entry_hours = tk.Entry(parent)
        self.entry_hours.grid(row=4, column=1, padx=5, pady=2, sticky="ew")
        self.check_fruehstueck = tk.IntVar()
        check_fruehstueck = tk.Checkbutton(
            parent,
            text="Frühstück (+0.25h)",
            variable=self.check_fruehstueck,
            command=self.toggle_fruehstueck,
        )
        check_fruehstueck.grid(row=5, column=1, sticky="w", pady=2, padx=5)
        self.check_mittagspause = tk.IntVar()
        check_mittagspause = tk.Checkbutton(
            parent,
            text="Mittagspause (+0.5h)",
            variable=self.check_mittagspause,
            command=self.toggle_mittagspause,
        )
        check_mittagspause.grid(row=6, column=1, sticky="w", pady=2, padx=5)
        self.check_urlaub = tk.IntVar()
        check_urlaub = tk.Checkbutton(
            parent,
            text="Urlaub",
            variable=self.check_urlaub,
            command=self.toggle_urlaub,
        )
        check_urlaub.grid(row=7, column=1, sticky="w", pady=2, padx=5)
        self.check_krank = tk.IntVar()
        check_krank = tk.Checkbutton(
            parent, text="Krank", variable=self.check_krank, command=self.toggle_krank
        )
        check_krank.grid(row=8, column=1, sticky="w", pady=2, padx=5)

        self.check_skug = tk.IntVar()
        check_skug = tk.Checkbutton(
            parent, text="No SKUG", variable=self.check_skug, command=self.toggle_skug
        )
        check_skug.grid(row=9, column=1, sticky="w", pady=2, padx=5)

        travel_frame = tk.Frame(parent)
        travel_frame.grid(row=10, column=1, sticky="w", pady=2, padx=5)

        self.check_reise = tk.IntVar()
        check_reise = tk.Checkbutton(
            travel_frame,
            text="Reise",
            variable=self.check_reise,
            command=self.toggle_reise,
        )
        check_reise.pack(side=tk.LEFT)

        self.combo_reise_type = ttk.Combobox(travel_frame, width=10, state="disabled")
        self.combo_reise_type["values"] = [ts.value for ts in TravelStatus]
        self.combo_reise_type.current(0)
        self.combo_reise_type.pack(side=tk.LEFT, padx=5)

        tk.Label(parent, text="Baustelle:").grid(
            row=11, column=0, sticky="e", padx=5, pady=2
        )
        bst_frame = tk.Frame(parent)
        bst_frame.grid(row=11, column=1, padx=5, pady=2, sticky="ew")
        self.entry_bst = tk.Entry(bst_frame)
        self.entry_bst.pack(side=tk.LEFT, fill=tk.X, expand=True)
        btn_bst_manager = tk.Button(
            bst_frame, text="⚙", width=2, command=self.open_baustelle_manager
        )
        btn_bst_manager.pack(side=tk.LEFT, padx=(2, 0))

        self.check_delete_mode = tk.IntVar()
        check_delete = tk.Checkbutton(
            parent, text="Entfernt Checkbox", variable=self.check_delete_mode, fg="red"
        )
        check_delete.grid(row=12, column=1, sticky="w", pady=2, padx=5)

        btn_frame = tk.Frame(parent)
        btn_frame.grid(row=13, column=0, columnspan=2, pady=20)

        btn_submit = tk.Button(btn_frame, text="Speichern", command=self.submit)
        btn_submit.pack(side=tk.LEFT, padx=5)

        btn_export = tk.Button(
            btn_frame, text="Excel Export", command=self.export_excel
        )
        btn_export.pack(side=tk.LEFT, padx=5)

        btn_preview = tk.Button(
            btn_frame, text="Excel Vorschau", command=self.open_excel_preview
        )
        btn_preview.pack(side=tk.LEFT, padx=5)

        btn_settings = tk.Button(
            btn_frame, text="⚙ Einstellungen", command=self.open_settings
        )
        btn_settings.pack(side=tk.LEFT, padx=5)

        self.edit_status_label = tk.Label(btn_frame, text="Bearbeiten: AUS", fg="gray")
        self.edit_status_label.pack(side=tk.LEFT, padx=10)
        self.btn_abort_edit = tk.Button(
            btn_frame, text="Abbrechen", command=self.clear_edit_mode, state="disabled"
        )
        self.btn_abort_edit.pack(side=tk.LEFT, padx=5)

        parent.grid_columnconfigure(1, weight=1)

        self.fields = [
            self.entry_year,
            self.entry_month,
            self.entry_day,
            self.entry_name,
            self.entry_hours,
            self.entry_bst,
        ]

        self.setup_autocomplete()

    def toggle_reise(self):
        if self.check_reise.get():
            self.combo_reise_type.config(state="readonly")
            if self.check_krank.get():
                self.clear_krank()
            if self.check_urlaub.get():
                self.clear_urlaub()
            self.entry_hours.config(state="normal")
        else:
            self.combo_reise_type.config(state="disabled")

    def toggle_fruehstueck(self):
        if self.check_fruehstueck.get():
            if self.check_krank.get():
                self.clear_krank()
            if self.check_urlaub.get():
                self.clear_urlaub()
            self.entry_hours.config(state="normal")

    def toggle_mittagspause(self):
        if self.check_mittagspause.get():
            if self.check_krank.get():
                self.clear_krank()
            if self.check_urlaub.get():
                self.clear_urlaub()
            self.entry_hours.config(state="normal")

    def toggle_skug(self):
        pass

    def create_data_displays(self, parent):
        display_paned = tk.PanedWindow(
            parent, orient=tk.VERTICAL, sashwidth=5, sashrelief=tk.RAISED, bg="gray"
        )
        display_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        month_frame = tk.LabelFrame(
            display_paned, text="Monat Übersicht (Jahr/Monat/Name(n))", padx=5, pady=5
        )
        display_paned.add(month_frame, minsize=200, stretch="always")

        month_columns = (
            "Tag",
            "Wochentag",
            "Name",
            "Kostenstelle",
            "Stunden",
            "F",
            "M",
            "SKUG",
            "Reise",
            "≤ 8h",
            "Löschen",
        )
        self.month_tree = ttk.Treeview(
            month_frame, columns=month_columns, show="headings", height=8
        )

        self.month_sort_column = "Tag"
        self.month_sort_reverse = False

        for col in month_columns:
            if col == "Löschen":
                self.month_tree.heading(col, text=col)
                self.month_tree.column(col, width=60, anchor="center")
            else:
                self.month_tree.heading(
                    col, text=col, command=lambda c=col: self.sort_month_tree(c)
                )
                if col == "Tag":
                    self.month_tree.column(col, width=50, anchor="center")
                elif col == "Wochentag":
                    self.month_tree.column(col, width=80, anchor="center")
                elif col == "Name":
                    self.month_tree.column(col, width=100, anchor="center")
                elif col == "Reise":
                    self.month_tree.column(col, width=80, anchor="center")
                elif col in ("F", "M"):
                    self.month_tree.column(col, width=30, anchor="center")
                else:
                    self.month_tree.column(col, width=80, anchor="center")

        self.month_tree.tag_configure("row_red", background="#FF9999")
        self.month_tree.tag_configure("row_even", background="#E0E0E0")
        self.month_tree.tag_configure("row_odd", background="#FFFFFF")
        month_scrollbar = ttk.Scrollbar(
            month_frame, orient=tk.VERTICAL, command=self.month_tree.yview
        )
        self.month_tree.configure(yscrollcommand=month_scrollbar.set)

        self.month_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        month_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.month_tree.bind("<ButtonRelease-1>", self.on_month_tree_click)

        day_frame = tk.LabelFrame(
            display_paned,
            text="Tages Übersicht (Jahr/Monat/Tag(e)/Baustelle)",
            padx=5,
            pady=5,
        )
        display_paned.add(day_frame, minsize=200, stretch="always")

        day_columns = ("Tag", "Wochentag", "Name", "Stunden", "SKUG", "Reise", "≤ 8h")
        self.day_tree = ttk.Treeview(
            day_frame, columns=day_columns, show="headings", height=8
        )

        self.day_sort_column = "Tag"
        self.day_sort_reverse = False

        for col in day_columns:
            self.day_tree.heading(
                col, text=col, command=lambda c=col: self.sort_day_tree(c)
            )
            if col == "Tag":
                self.day_tree.column(col, width=50, anchor="center")
            elif col == "Wochentag":
                self.day_tree.column(col, width=80, anchor="center")
            elif col == "Name":
                self.day_tree.column(col, width=100, anchor="center")
            elif col == "Reise":
                self.day_tree.column(col, width=80, anchor="center")
            else:
                self.day_tree.column(col, width=90, anchor="center")

        self.day_tree.tag_configure("row_even", background="#E0E0E0")
        self.day_tree.tag_configure("row_odd", background="#FFFFFF")

        day_scrollbar = ttk.Scrollbar(
            day_frame, orient=tk.VERTICAL, command=self.day_tree.yview
        )
        self.day_tree.configure(yscrollcommand=day_scrollbar.set)

        self.day_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        day_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_bindings(self):
        self.entry_year.bind("<KeyRelease>", self.update_weekday)
        self.entry_month.bind("<KeyRelease>", self.update_weekday)
        self.entry_day.bind("<KeyRelease>", self.update_weekday)

        self.entry_year.bind("<KeyRelease>", self.update_month_view, add="+")
        self.entry_month.bind("<KeyRelease>", self.update_month_view, add="+")
        self.entry_name.bind("<KeyRelease>", self.update_month_view, add="+")

        self.entry_year.bind("<KeyRelease>", self.update_day_view, add="+")
        self.entry_month.bind("<KeyRelease>", self.update_day_view, add="+")
        self.entry_day.bind("<KeyRelease>", self.update_day_view, add="+")
        self.entry_bst.bind("<KeyRelease>", self.update_day_view, add="+")

        self.entry_year.bind("<KeyRelease>", self.check_edit_mode_abort, add="+")
        self.entry_month.bind("<KeyRelease>", self.check_edit_mode_abort, add="+")
        self.entry_day.bind("<KeyRelease>", self.check_edit_mode_abort, add="+")
        self.entry_name.bind("<KeyRelease>", self.check_edit_mode_abort, add="+")

        self.entry_year.bind("<KeyRelease>", self.schedule_preview_refresh, add="+")
        self.entry_month.bind("<KeyRelease>", self.schedule_preview_refresh, add="+")

        autocomplete_fields = [self.entry_name, self.entry_bst]
        for field in self.fields:
            if field not in autocomplete_fields:
                field.bind("<Return>", self.focus_next)
                field.bind("<Down>", self.focus_next)
                field.bind("<Up>", self.focus_previous)
            else:
                field.bind("<Return>", self.focus_next, add="+")
                field.bind("<Down>", self.focus_next, add="+")
                field.bind("<Up>", self.focus_previous, add="+")

        # Bind autocomplete selection events to update views
        self.entry_name.bind(
            "<<AutocompleteSelected>>", self.update_month_view, add="+"
        )
        self.entry_bst.bind("<<AutocompleteSelected>>", self.update_day_view, add="+")
        self.entry_name.bind(
            "<<AutocompleteSelected>>", self.check_edit_mode_abort, add="+"
        )

        self.month_tree.bind("<Double-1>", self.load_entry_from_tree)
        self.day_tree.bind("<Double-1>", self.load_entry_from_tree)

        # Bind click event to root to detect clicks outside of treeviews
        self.root.bind("<Button-1>", self.on_global_click, add="+")

    def update_weekday(self, *args):
        tag_input = self.entry_day.get().strip()
        jahr = self.entry_year.get()
        monat = self.entry_month.get()

        skip_weekends = self.settings.get("skip_weekends", True)
        skip_holidays = self.settings.get("skip_holidays", True)

        try:
            jahr_int = int(jahr) if jahr else None
            monat_int = int(monat) if monat else None

            if jahr_int and monat_int:
                days = parse_date_range(
                    tag_input, jahr_int, monat_int, skip_weekends, skip_holidays
                )
            else:
                days = parse_date_range(tag_input)
        except (ValueError, TypeError):
            days = parse_date_range(tag_input)

        if days:
            if len(days) == 1:
                weekday = get_weekday_abbr(jahr, monat, str(days[0]))
                if weekday:
                    self.label_day.config(text=f"Tag(e) ({weekday}):*")
                else:
                    self.label_day.config(text="Tag(e):*")
            else:
                self.label_day.config(text=f"Tag(e) ({len(days)} Tage):*")
        else:
            try:
                single_day = int(tag_input)
                weekday = get_weekday_abbr(jahr, monat, str(single_day))
                if weekday:
                    self.label_day.config(text=f"Tag(e) ({weekday}):*")
                else:
                    self.label_day.config(text="Tag(e):*")
            except (ValueError, TypeError):
                self.label_day.config(text="Tag(e):*")

    def update_month_view(self, *args):
        for item in self.month_tree.get_children():
            self.month_tree.delete(item)

        year = self.entry_year.get().strip()
        month = self.entry_month.get().strip()
        names_input = self.entry_name.get().strip()

        if not (year and month):
            return

        try:
            year_int = int(year)
            month_int = int(month)

            names = parse_multiple_names(names_input)

            if not names:
                return

            all_entries = []
            for name in names:
                entries = self.db.get_arbeitsstunden_for_month(
                    year_int, month_int, name
                )
                all_entries.extend(entries)

            all_entries.sort(key=lambda x: x["tag"])

            for i, entry in enumerate(all_entries):
                tags = []
                print(entry)
                meta_data = (
                    self.db.get_metadata_by_date(
                        year_int, month_int, entry["tag"], entry["name"]
                    )
                    or {}
                )
                if meta_data.get("kg_8h"):
                    tags.append("row_red")
                else:
                    tags.append("row_even" if i % 2 == 0 else "row_odd")

                tags.append(f"entry_{entry['id']}")

                self.month_tree.insert(
                    "",
                    tk.END,
                    values=(
                        entry["tag"],
                        entry["wochentag"] or "",
                        entry["name"],
                        entry["kostenstelle"] or "",
                        entry["stunden"] or "",
                        "X" if meta_data.get("fruehstueck") else "",
                        "X" if meta_data.get("mittag") else "",
                        meta_data.get("skug") or "",
                        meta_data.get("travel_status") or "",
                        "Ja"
                        if meta_data.get("kg_8h")
                        else ("" if meta_data.get("kg_8h") is None else "Nein"),
                        "🗑",
                    ),
                    tags=tuple(tags),
                )

        except (ValueError, TypeError):
            pass

    def update_day_view(self, *args):
        for item in self.day_tree.get_children():
            self.day_tree.delete(item)

        year = self.entry_year.get().strip()
        month = self.entry_month.get().strip()
        day_input = self.entry_day.get().strip()
        baustelle = self.entry_bst.get().strip()

        if not (year and month and day_input and baustelle):
            return

        try:
            year_int = int(year)
            month_int = int(month)

            skip_weekends = self.settings.get("skip_weekends", True)
            skip_holidays = self.settings.get("skip_holidays", True)

            days = parse_date_range(
                day_input, year_int, month_int, skip_weekends, skip_holidays
            )

            if days is None:
                try:
                    single_day = int(day_input)
                    if 1 <= single_day <= 31:
                        days = [single_day]
                    else:
                        return
                except ValueError:
                    return

            all_entries = []
            for day in days:
                entries = self.db.get_entries_by_date_and_baustelle(
                    year_int, month_int, day, baustelle
                )
                all_entries.extend(entries)

            all_entries.sort(key=lambda x: x["tag"])

            for i, entry in enumerate(all_entries):
                wochentag = (
                    get_weekday_abbr(str(year_int), str(month_int), str(entry["tag"]))
                    or ""
                )
                meta_data = (
                    self.db.get_metadata_by_date(
                        year_int, month_int, entry["tag"], entry["name"]
                    )
                    or {}
                )
                row_tag = "row_even" if i % 2 == 0 else "row_odd"
                self.day_tree.insert(
                    "",
                    tk.END,
                    values=(
                        entry["tag"],
                        wochentag,
                        entry["name"],
                        entry["stunden"] or "",
                        meta_data.get("skug") or "",
                        meta_data.get("travel_status") or "",
                        "Ja"
                        if meta_data.get("kg_8h")
                        else ("" if meta_data.get("kg_8h") is None else "Nein"),
                    ),
                    tags=(row_tag, f"entry_{entry['id']}"),
                )

        except (ValueError, TypeError):
            pass

    def on_month_tree_click(self, event):
        region = self.month_tree.identify_region(event.x, event.y)

        # If click is not on a cell (e.g. empty space), deselect
        if region != "cell":
            self.month_tree.selection_remove(self.month_tree.selection())
            return

        column = self.month_tree.identify_column(event.x)

        if column == "#11":
            item = self.month_tree.identify_row(event.y)
            if item:
                tags = self.month_tree.item(item, "tags")
                if tags:
                    entry_id_str = tags[1]  # Format: "entry_123"
                    entry_id = int(entry_id_str.split("_")[1])

                    values = self.month_tree.item(item, "values")
                    name = values[2]
                    tag = values[0]

                    if messagebox.askyesno(
                        "Eintrag löschen",
                        f"Möchten Sie den Eintrag für {name} am Tag {tag} wirklich löschen?",
                    ):
                        if self.db.delete_arbeitsstunden(entry_id):
                            self.month_tree.delete(item)
                            self.update_day_view()
                            self.schedule_preview_refresh()
                        else:
                            messagebox.showerror(
                                "Fehler", "Eintrag konnte nicht gelöscht werden."
                            )

    def on_global_click(self, event):
        # Check if the click happened outside of the treeviews
        widget = event.widget
        # If the widget is not one of the treeviews, deselect items
        if widget != self.month_tree and widget != self.day_tree:
            # Also check if it's not a scrollbar of the treeview (optional, but good practice)
            # Simplest approach: if widget is not the treeview itself.

            # Note: Clicking ON the treeview logic is handled by treeview bindings usually.
            # But if we click on "background" of main window, we want to deselect.

            # self.month_tree.selection() returns a tuple of selected items
            if self.month_tree.selection():
                self.month_tree.selection_remove(self.month_tree.selection())

            if self.day_tree.selection():
                self.day_tree.selection_remove(self.day_tree.selection())

    def load_entry_from_tree(self, event):
        tree = event.widget
        item_id = tree.identify_row(event.y)
        if not item_id:
            return

        tags = tree.item(item_id, "tags")
        entry_id = self.get_entry_id_from_tags(tags)
        entry_data = self.db.get_arbeitsstunden_by_id(entry_id) if entry_id else None

        values = tree.item(item_id, "values")
        if entry_data:
            tag = entry_data.get("tag")
            name = entry_data.get("name")
            stunden = entry_data.get("stunden")
            bst = entry_data.get("kostenstelle")
            jahr = entry_data.get("jahr")
            monat = entry_data.get("monat")
        else:
            tag = values[0]
            name = values[2]
            stunden = ""
            bst = ""
            jahr = self.entry_year.get().strip()
            monat = self.entry_month.get().strip()
            if tree == self.month_tree:
                bst = values[3]
                stunden = values[4]
            elif tree == self.day_tree:
                stunden = values[3]
                bst = self.entry_bst.get()

        self.entry_day.delete(0, tk.END)
        self.entry_day.insert(0, str(tag))

        self.entry_name.delete(0, tk.END)
        self.entry_name.insert(0, name)

        self.entry_hours.delete(0, tk.END)
        self.entry_hours.insert(0, str(stunden) if stunden else "")

        self.entry_bst.delete(0, tk.END)
        self.entry_bst.insert(0, str(bst) if bst else "")

        self.update_weekday()
        self.update_day_view()
        self.update_month_view()

        if entry_id and jahr and monat and tag and name:
            self.set_edit_mode(entry_id, jahr, monat, tag, name)
            self.load_metadata_flags(jahr, monat, tag, name)
        # Trigger updates?
        # Changing Day/Name/BST triggers auto updates via bindings on KeyRelease.
        # `insert` does NOT trigger KeyRelease. We might need to manually trigger update if we want views to filter?
        # But usually double click is to EDIT. So we just fill fields.
        # User can then modify and click "Speichern".

    def toggle_krank(self):
        self.entry_hours.delete(0, tk.END)
        self.entry_hours.insert(0, "0")
        if self.check_krank.get():
            self.check_urlaub.set(0)
            self.check_fruehstueck.set(0)
            self.check_mittagspause.set(0)
            self.check_skug.set(0)
            self.check_reise.set(0)
            self.combo_reise_type.set(TravelStatus.Nicht)
            self.entry_hours.config(state="disabled")
            self.entry_bst.config(state="normal")
            self.entry_bst.delete(0, tk.END)
            self.entry_bst.insert(0, "Krank")
            self.entry_bst.config(state="disabled")
        else:
            self.clear_krank()

    def clear_krank(self):
        self.check_krank.set(0)
        self.entry_hours.config(state="normal")
        self.entry_hours.delete(0, tk.END)
        self.entry_bst.config(state="normal")
        self.entry_bst.delete(0, tk.END)

    def toggle_urlaub(self):
        self.entry_hours.delete(0, tk.END)
        self.entry_hours.insert(0, "0")
        if self.check_urlaub.get():
            self.check_krank.set(0)
            self.check_fruehstueck.set(0)
            self.check_mittagspause.set(0)
            self.check_skug.set(0)
            self.check_reise.set(0)
            self.combo_reise_type.set(TravelStatus.Nicht)
            self.entry_hours.config(state="disabled")
            self.entry_bst.config(state="normal")
            self.entry_bst.delete(0, tk.END)
            self.entry_bst.insert(0, self.get_urlaub_kostenstelle())
            self.entry_bst.config(state="disabled")
        else:
            self.clear_urlaub()

    def clear_urlaub(self):
        self.check_urlaub.set(0)
        self.entry_hours.config(state="normal")
        self.entry_hours.delete(0, tk.END)
        self.entry_bst.config(state="normal")
        self.entry_bst.delete(0, tk.END)

    def submit(self):
        # handles Jahr, Monat, Name, Stunden(, -> .)
        jahr_input = self.entry_year.get().strip()
        monat_input = self.entry_month.get().strip()
        tag_input = self.entry_day.get().strip()
        names_input = self.entry_name.get().strip()

        stunden_input = self.entry_hours.get().strip()
        stunden_input = stunden_input.replace(",", ".")
        new_stunden = float(stunden_input) if stunden_input else None

        baustelle_input = self.entry_bst.get().strip()

        input_fruehstueck = bool(self.check_fruehstueck.get())
        input_mittag = bool(self.check_mittagspause.get())
        input_urlaub = bool(self.check_urlaub.get())
        input_krank = bool(self.check_krank.get())
        input_no_skug = bool(self.check_skug.get())
        input_reise = bool(self.check_reise.get())
        delete_mode = bool(self.check_delete_mode.get())

        edit_mode_for_submit = self.edit_mode_active
        if edit_mode_for_submit:
            if (
                jahr_input != (self.edit_original_year or "")
                or monat_input != (self.edit_original_month or "")
                or tag_input != (self.edit_original_day or "")
                or names_input != (self.edit_original_name or "")
            ):
                self.clear_edit_mode()
                edit_mode_for_submit = False

        travel_type_input = self.combo_reise_type.get()
        if travel_type_input == TravelStatus.Nicht:
            travel_type_input = None

        is_valid, error_msg = validate_required_fields(
            jahr_input, monat_input, names_input, stunden_input
        )

        if not is_valid:
            messagebox.showerror("Validierungsfehler", error_msg)
            return

        names = parse_multiple_names(names_input)

        if not names:
            messagebox.showerror("Fehler", "Mindestens ein Name muss angegeben werden!")
            return

        try:
            jahr_int = int(jahr_input)
            monat_int = int(monat_input)
        except ValueError:
            messagebox.showerror("Fehler", "Ungültiges Jahr oder Monat!")
            return

        skip_weekends = self.settings.get("skip_weekends", True)
        skip_holidays = self.settings.get("skip_holidays", True)

        days = parse_date_range(
            tag_input, jahr_int, monat_int, skip_weekends, skip_holidays
        )

        if days is None:
            try:
                single_day = int(tag_input)
                if 1 <= single_day <= 31:
                    days = [single_day]
                else:
                    messagebox.showerror("Fehler", "Tag muss zwischen 1 und 31 liegen!")
                    return
            except ValueError:
                messagebox.showerror(
                    "Fehler",
                    "Ungültiges Tag-Format! Verwenden Sie z.B. '3-7,9,11-13' oder '5'",
                )
                return

        if not days:
            messagebox.showwarning(
                "Warnung",
                "Alle eingegebenen Tage wurden gefiltert (Wochenenden/Feiertage).\n"
                "Bitte passen Sie die Einstellungen an oder wählen Sie andere Tage.",
            )
            return

        is_valid, invalid_days = validate_days_in_month(jahr_int, monat_int, days)
        if not is_valid:
            invalid_days_str = ", ".join(map(str, invalid_days))
            messagebox.showerror(
                "Fehler",
                f"Die folgenden Tage existieren nicht im Monat {monat_input}/{jahr_input}:\n{invalid_days_str}",
            )
            return

        if new_stunden is None and not any(
            [
                input_fruehstueck,
                input_mittag,
                input_urlaub,
                input_krank,
                input_no_skug,
                input_reise,
                delete_mode,
            ]
        ):
            messagebox.showinfo(
                "Info", "Keine Stunden und keine Optionen gewählt - nichts zu tun."
            )
            return

        if not input_krank and not input_urlaub:
            if baustelle_input and not self.entry_service.is_valid_kostenstelle(
                baustelle_input
            ):
                messagebox.showerror(
                    "Fehler",
                    "Ungültige Kostenstelle. Bitte eine Baustelle auswählen oder 'Krank', '900', '940' verwenden.",
                )
                return

        skug_settings = self.master_db.get_skug_settings()

        edit_entry_data = None
        if edit_mode_for_submit:
            if len(names) != 1 or len(days) != 1:
                self.clear_edit_mode()
                edit_mode_for_submit = False
            else:
                edit_entry_data = self.db.get_arbeitsstunden_by_id(self.edit_entry_id)
                if not edit_entry_data:
                    self.clear_edit_mode()
                    edit_mode_for_submit = False
                else:
                    if (
                        str(edit_entry_data.get("jahr")) != jahr_input
                        or str(edit_entry_data.get("monat")) != monat_input
                        or str(edit_entry_data.get("tag")) != str(days[0])
                        or edit_entry_data.get("name") != names[0]
                    ):
                        self.clear_edit_mode()
                        edit_mode_for_submit = False

        total_entries = 0
        updated_entries = 0
        errors = []
        sorted_days = sorted(days)
        wants_day_metadata = input_fruehstueck or input_mittag or input_reise
        try:
            if (
                new_stunden is not None
                and not input_krank
                and not input_urlaub
                and not baustelle_input
            ):
                for name in names:
                    for day in sorted_days:
                        existing_entries = self.db.get_arbeitsstunden_for_day(
                            jahr_int, monat_int, day, name
                        )
                        if any(
                            e.get("kostenstelle") in ["Krank", "900", "940"]
                            for e in existing_entries
                        ):
                            messagebox.showerror(
                                "Fehler",
                                "Stunden ohne Kostenstelle sind nicht erlaubt, wenn bereits Krank/Urlaub erfasst ist.",
                            )
                            return

            for name in names:
                for i, day in enumerate(sorted_days):
                    if input_krank or input_urlaub:
                        handle_krank_urlaub(
                            jahr_int,
                            monat_int,
                            day,
                            name,
                            self.db,
                            self.master_db,
                            input_krank,
                            input_urlaub,
                            skug_settings,
                        )
                        total_entries += 1
                        continue

                    if (
                        edit_mode_for_submit
                        and edit_entry_data
                        and name == edit_entry_data.get("name")
                        and day == int(edit_entry_data.get("tag"))
                    ):
                        target_entry_id = self.edit_entry_id
                        entry_data = dict(edit_entry_data)
                        errors = []
                    else:
                        target_entry_id, entry_data, errors = try_load_existing_entry(
                            jahr_int, monat_int, day, name, baustelle_input, self.db
                        )
                    metadata_entry = self.db.get_metadata_by_date(
                        jahr_int, monat_int, day, name
                    )
                    if not metadata_entry:
                        metadata_entry = {}
                        wochentag = (
                            get_weekday_abbr(jahr_int, monat_int, str(day)) or ""
                        )
                        metadata_entry.update(
                            {
                                "jahr": jahr_int,
                                "monat": monat_int,
                                "tag": str(day),
                                "name": name,
                                "wochentag": wochentag,
                            }
                        )

                    if errors:
                        continue

                    has_existing_work_entry = self.day_has_work_entry(
                        jahr_int, monat_int, day, name
                    )
                    creates_work_entry_now = new_stunden is not None and bool(
                        baustelle_input or entry_data.get("Kostenstelle")
                    )
                    if wants_day_metadata and not (
                        has_existing_work_entry or creates_work_entry_now
                    ):
                        errors.append(
                            f"Fruehstueck, Mittag und Reise sind nur mit einem Arbeitseintrag erlaubt: {name} am {day}.{monat_int}.{jahr_int}."
                        )
                        continue

                    if new_stunden is not None:
                        entry_data["Stunden"] = new_stunden

                    if baustelle_input:
                        entry_data["Kostenstelle"] = baustelle_input
                    elif (
                        new_stunden is not None and not input_krank and not input_urlaub
                    ):
                        if not self.entry_service.is_valid_kostenstelle(
                            entry_data.get("Kostenstelle")
                        ):
                            errors.append(
                                "Kostenstelle fehlt oder ist ungültig. Bitte Baustelle auswählen."
                            )
                            continue

                    if not target_entry_id:
                        wochentag = (
                            get_weekday_abbr(jahr_int, monat_int, str(day)) or ""
                        )
                        entry_data.update(
                            {
                                "jahr": jahr_int,
                                "monat": monat_int,
                                "tag": str(day),
                                "name": name,
                                "wochentag": wochentag,
                                "stunden": new_stunden
                                if new_stunden is not None
                                else 0.0,
                                "kostenstelle": baustelle_input,
                            }
                        )

                    if input_fruehstueck:
                        metadata_entry["fruehstueck"] = True
                    if input_mittag:
                        metadata_entry["mittag"] = True

                    if edit_mode_for_submit:
                        metadata_entry["fruehstueck"] = input_fruehstueck
                        metadata_entry["mittag"] = input_mittag
                        metadata_entry["urlaub"] = None
                        metadata_entry["krank"] = None

                    if input_reise:
                        final_travel_status = None
                        if travel_type_input == TravelStatus.Auto:
                            if len(sorted_days) == 1:
                                final_travel_status = "Anreise"
                            else:
                                if i == 0:
                                    final_travel_status = "Anreise"
                                elif i == len(sorted_days) - 1:
                                    final_travel_status = "Abreise"
                                else:
                                    final_travel_status = "24h_away"
                        else:
                            final_travel_status = travel_type_input
                        metadata_entry["travel_status"] = final_travel_status
                    elif edit_mode_for_submit:
                        metadata_entry["travel_status"] = None

                    metadata_entry["no_skug"] = input_no_skug

                    self.db.add_or_update_metadata(metadata_entry)
                    if not check_arbeitsstunden(entry_data):
                        pass
                    elif target_entry_id:
                        print("Update arbeitsentry")
                        self.db.update_arbeitsstunden(target_entry_id, entry_data)
                    elif new_stunden is None:
                        pass
                    else:
                        print("Add new arbeitsentry")
                        self.db.add_arbeitsstunden(entry_data)

                    if delete_mode and not edit_mode_for_submit:
                        if input_fruehstueck:
                            metadata_entry["fruehstueck"] = False
                        if input_mittag:
                            metadata_entry["mittag"] = False
                        if input_no_skug:
                            metadata_entry["no_skug"] = False
                        if input_reise:
                            metadata_entry["travel_status"] = None

                    is_winter = monat_int in [12, 1, 2, 3]
                    if is_winter and not metadata_entry.get("no_skug", False):
                        arbeits_stunden = sum(
                            [
                                entry["stunden"]
                                for entry in self.db.get_arbeitsstunden_for_day(
                                    jahr_int, monat_int, day, name
                                )
                            ]
                        )
                        skug = calculate_skug(
                            jahr_int, monat_int, day, arbeits_stunden, skug_settings
                        )
                        metadata_entry["skug"] = skug if skug >= 1 else 0
                    else:
                        metadata_entry["skug"] = None

                    total_entries += 1

                    is_unter_8h = determine_kg_8h_flag(
                        self.db, self.master_db, jahr_int, monat_int, day, name
                    )
                    metadata_entry["kg_8h"] = is_unter_8h
                    self.db.add_or_update_metadata(metadata_entry)

            if errors:
                error_msg = (
                    f"{total_entries} Einträge verarbeitet.\n\nFehler:\n"
                    + "\n".join(errors[:10])
                )
                if len(errors) > 10:
                    error_msg += f"\n... und {len(errors) - 10} weitere Fehler"
                messagebox.showwarning("Hinweis", error_msg)

        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern:\n{str(e)}")
            print(e)

        self.update_month_view()
        self.update_day_view()
        self.schedule_preview_refresh()

        if self.settings.get("auto_increment_day", False) and not delete_mode:
            last_day = max(days)

            if self.settings.get("skip_weekends", True):
                next_year, next_month, next_day = get_next_day_skip_weekend(
                    jahr_int, monat_int, last_day
                )
            else:
                next_year, next_month, next_day = get_next_day(
                    jahr_int, monat_int, last_day
                )

            if next_year != jahr_int:
                self.entry_year.delete(0, tk.END)
                self.entry_year.insert(0, str(next_year))
            if next_month != monat_int:
                self.entry_month.delete(0, tk.END)
                self.entry_month.insert(0, str(next_month))

            self.entry_day.delete(0, tk.END)
            self.entry_day.insert(0, str(next_day))

            self.update_weekday()

        should_clear_baustelle = False
        if input_krank or input_urlaub:
            should_clear_baustelle = True

        self.clear_fields(clear_baustelle=should_clear_baustelle)

        cursor_target = self.settings.get("cursor_jump_target", "Tag")
        if cursor_target == "Tag":
            self.entry_day.focus()
        elif cursor_target == "Name":
            self.entry_name.focus()
        elif cursor_target == "Stunden":
            self.entry_hours.focus()
        elif cursor_target == "Baustelle":
            self.entry_bst.focus()
        else:
            self.entry_day.focus()

    def export_excel(self):
        try:
            jahr_str = self.entry_year.get().strip()
            monat_str = self.entry_month.get().strip()

            if not jahr_str or not monat_str:
                messagebox.showwarning("Warnung", "Bitte Jahr und Monat eingeben.")
                return

            try:
                jahr = int(jahr_str)
                monat = int(monat_str)

                if monat < 1 or monat > 12:
                    messagebox.showwarning(
                        "Warnung", "Monat muss zwischen 1 und 12 liegen."
                    )
                    return

            except ValueError:
                messagebox.showwarning(
                    "Warnung", "Jahr und Monat müssen gültige Zahlen sein."
                )
                return

            if export_to_excel_top_to_bottom(jahr, monat, self.db, self.master_db):
                messagebox.showinfo(
                    "Erfolg", f"Daten für {monat:02d}/{jahr} nach Excel exportiert!"
                )
            else:
                messagebox.showwarning(
                    "Warnung", "Keine Daten zum Exportieren vorhanden."
                )
        except Exception as e:
            messagebox.showerror("Fehler", f"Export fehlgeschlagen:\n{str(e)}")

    def open_excel_preview(self):
        if self.preview_window is None or not self.preview_window.is_open():
            self.preview_window = ExcelPreviewWindow(
                self.root,
                on_close=self.clear_preview_window,
                on_edit=self.handle_preview_edit,
                on_apply=self.apply_preview_changes,
                on_flag=self.handle_preview_flag,
                on_reset=self.reset_preview_changes,
                on_modified=self.sync_preview_pending_edits_from_sheet,
            )
        else:
            self.preview_window.window.lift()
            self.preview_window.window.focus_force()

        self.refresh_preview_from_entries()

    def clear_preview_window(self):
        self.preview_window = None
        if self.preview_refresh_job is not None:
            self.root.after_cancel(self.preview_refresh_job)
            self.preview_refresh_job = None
        self.preview_pending_request = None
        if self.preview_task_future is not None and not self.preview_task_future.done():
            self.preview_task_future.cancel()
        self.preview_pending_edits = {}
        self.preview_pending_flags = {}

    def on_app_close(self):
        self.shutdown_preview_executor()
        self.root.destroy()

    def shutdown_preview_executor(self):
        if self.preview_executor is None:
            return
        try:
            self.preview_executor.shutdown(wait=False, cancel_futures=True)
        except TypeError:
            self.preview_executor.shutdown(wait=False)
        self.preview_executor = None

    def schedule_preview_refresh(self, event=None):
        if self.preview_window is None or not self.preview_window.is_open():
            return
        if self.preview_pending_edits or self.preview_pending_flags:
            return
        if self.preview_refresh_job is not None:
            self.root.after_cancel(self.preview_refresh_job)
        self.preview_refresh_job = self.root.after(
            300, self.refresh_preview_from_entries
        )

    def refresh_preview_from_entries(self, force=False):
        if self.preview_window is None or not self.preview_window.is_open():
            return
        if not force and (self.preview_pending_edits or self.preview_pending_flags):
            return
        self.preview_refresh_job = None

        year = self.entry_year.get().strip()
        month = self.entry_month.get().strip()

        if not year or not month:
            self.preview_window.show_message(
                "Excel Vorschau", "Bitte Jahr und Monat eingeben."
            )
            return

        try:
            year_int = int(year)
            month_int = int(month)
        except ValueError:
            self.preview_window.show_message(
                "Excel Vorschau", "Jahr und Monat müssen gültige Zahlen sein."
            )
            return

        if month_int < 1 or month_int > 12:
            self.preview_window.show_message(
                "Excel Vorschau", "Monat muss zwischen 1 und 12 liegen."
            )
            return

        self.queue_preview_build(year_int, month_int)

    def queue_preview_build(self, year_int, month_int):
        if self.preview_window is None or not self.preview_window.is_open():
            return
        self.preview_request_seq += 1
        request_id = self.preview_request_seq

        if self.preview_task_future is not None and not self.preview_task_future.done():
            self.preview_pending_request = (request_id, year_int, month_int)
            self.preview_window.show_message(
                f"Excel Vorschau {month_int:02d}/{year_int}",
                "Vorschau wird aktualisiert...",
            )
            return

        self.start_preview_build(request_id, year_int, month_int)

    def start_preview_build(self, request_id, year_int, month_int):
        self.preview_inflight_request_id = request_id
        self.preview_window.show_message(
            f"Excel Vorschau {month_int:02d}/{year_int}",
            "Lade Vorschau...",
        )
        self.preview_task_future = self.preview_executor.submit(
            self.build_preview_workbook, year_int, month_int
        )
        self.preview_task_future.add_done_callback(
            lambda future: self.root.after(
                0,
                self.on_preview_ready,
                future,
                request_id,
                year_int,
                month_int,
            )
        )

    def build_preview_workbook(self, year_int, month_int):
        cell_map = {}
        workbook = build_workbook_top_to_bottom(
            year_int, month_int, self.db, self.master_db, cell_map
        )
        return workbook, cell_map

    def on_preview_ready(self, future, request_id, year_int, month_int):
        if self.preview_window is None or not self.preview_window.is_open():
            return
        if request_id != self.preview_inflight_request_id:
            return

        try:
            workbook, cell_map = future.result()
        except Exception as exc:
            self.preview_window.show_message(
                f"Excel Vorschau {month_int:02d}/{year_int}",
                f"Fehler beim Laden der Vorschau: {exc}",
            )
            workbook = None
            cell_map = None

        if workbook is not None:
            self.preview_window.render_workbook(
                workbook, year_int, month_int, cell_map=cell_map
            )

        if self.preview_pending_request is not None:
            pending_request_id, pending_year, pending_month = (
                self.preview_pending_request
            )
            self.preview_pending_request = None
            self.start_preview_build(pending_request_id, pending_year, pending_month)

    def handle_preview_edit(self, row, col, value):
        if self.preview_window is None or not self.preview_window.is_open():
            return
        if col is None or row is None:
            return
        if col <= 0:
            return

        wb_row = row + 1
        wb_col = col
        cell_info = self.preview_window.cell_map.get((wb_row, wb_col))
        if not cell_info:
            messagebox.showwarning("Hinweis", "Diese Zelle ist nicht editierbar.")
            self.schedule_preview_refresh()
            return

        raw_value = "" if value is None else str(value).strip()
        field = cell_info.get("field")
        year = cell_info.get("year")
        month = cell_info.get("month")
        day = cell_info.get("day")
        name = cell_info.get("name")
        entry_id = cell_info.get("entry_id")

        if not all([year, month, day, name]):
            messagebox.showerror("Fehler", "Ungültige Zellzuordnung.")
            self.schedule_preview_refresh()
            return

        original_value = self.preview_window.original_values.get((row, col))
        original_text = "" if original_value is None else str(original_value)
        is_changed = raw_value != original_text

        key = (row, col)
        if is_changed:
            self.preview_pending_edits[key] = {
                "cell_info": cell_info,
                "value": raw_value,
            }
        else:
            self.preview_pending_edits.pop(key, None)

        self.preview_window.mark_changed_cell(row, col, is_changed)
        self.preview_window.set_pending_count(
            len(self.preview_pending_edits) + len(self.preview_pending_flags)
        )

    def sync_preview_pending_edits_from_sheet(self):
        if self.preview_window is None or not self.preview_window.is_open():
            return
        if not self.preview_window.cell_map:
            return

        previous_edits = self.preview_pending_edits
        previous_keys = set(previous_edits.keys())
        new_edits = {}

        for (wb_row, wb_col), cell_info in self.preview_window.cell_map.items():
            row = wb_row - 1
            col = wb_col
            if row < 0 or col <= 0:
                continue
            flag_key = (
                cell_info.get("year"),
                cell_info.get("month"),
                cell_info.get("day"),
                cell_info.get("name"),
            )
            pending_day_flags = self.preview_pending_flags.get(flag_key, {})
            if cell_info.get("field") in ["Stunden", "Kostenstelle"] and (
                pending_day_flags.get("krank") or pending_day_flags.get("urlaub")
            ):
                continue
            current_value = self.preview_window.get_cell_text(row, col)
            raw_value = "" if current_value is None else str(current_value).strip()
            original_value = self.preview_window.original_values.get((row, col))
            original_text = "" if original_value is None else str(original_value)
            if raw_value != original_text:
                new_edits[(row, col)] = {
                    "cell_info": cell_info,
                    "value": raw_value,
                }

        self.preview_pending_edits = new_edits

        new_keys = set(new_edits.keys())
        removed_keys = previous_keys - new_keys
        for row, col in new_keys:
            self.preview_window.mark_changed_cell(row, col, True)

        for row, col in removed_keys:
            cell_info = None
            if (row, col) in previous_edits:
                cell_info = previous_edits[(row, col)].get("cell_info")
            if cell_info is None:
                cell_info = self.preview_window.cell_map.get((row + 1, col))
            keep_highlight = False
            if cell_info:
                flag_key = (
                    cell_info.get("year"),
                    cell_info.get("month"),
                    cell_info.get("day"),
                    cell_info.get("name"),
                )
                if flag_key in self.preview_pending_flags:
                    keep_highlight = True
            if not keep_highlight:
                self.preview_window.mark_changed_cell(row, col, False)

        self.preview_window.set_pending_count(
            len(self.preview_pending_edits) + len(self.preview_pending_flags)
        )

    def handle_preview_flag(
        self, year, month, day, name, flag, value, row=None, col=None
    ):
        key = (year, month, day, name)
        if key not in self.preview_pending_flags:
            self.preview_pending_flags[key] = {}
        self.preview_pending_flags[key][flag] = value

        if row is not None and col is not None:
            self.apply_preview_flag_visuals(
                year, month, day, name, flag, value, row, col
            )

        for (row, col), cell_info in self.preview_window.cell_map.items():
            if (
                cell_info.get("year") == year
                and cell_info.get("month") == month
                and cell_info.get("day") == day
                and cell_info.get("name") == name
            ):
                sheet_row = row - 1
                sheet_col = col
                self.preview_window.mark_changed_cell(sheet_row, sheet_col, True)

        self.preview_window.set_pending_count(
            len(self.preview_pending_edits) + len(self.preview_pending_flags)
        )

    def apply_preview_flag_visuals(self, year, month, day, name, flag, value, row, col):
        if self.preview_window is None or not self.preview_window.is_open():
            return
        if flag in ["krank", "urlaub"] and value:
            worker_type = (
                self.master_db.get_worker_type_by_name(name) or WorkerTypes.Fest
            )
            if flag == "krank":
                std_value = "K"
                bst_value = "900" if worker_type == WorkerTypes.Fest else "930"
            else:
                if worker_type == WorkerTypes.Fest:
                    std_value = "U"
                    bst_value = "900"
                else:
                    std_value = "F"
                    bst_value = "940"
            std_col = None
            bst_col = None
            for (wb_row, wb_col), info in self.preview_window.cell_map.items():
                if (
                    info.get("year") == year
                    and info.get("month") == month
                    and info.get("day") == day
                    and info.get("name") == name
                    and wb_row - 1 == row
                ):
                    if info.get("field") == "Stunden":
                        std_col = wb_col
                    elif info.get("field") == "Kostenstelle":
                        bst_col = wb_col

            if std_col is None:
                std_col = col
            if bst_col is None:
                bst_col = col + 1
            self.preview_window.set_cell_text(row, std_col, std_value)
            self.preview_window.set_cell_text(row, bst_col, bst_value)
            self.preview_window.mark_changed_cell(row, std_col, True)
            self.preview_window.mark_changed_cell(row, bst_col, True)

    def apply_preview_changes(self):
        self.sync_preview_pending_edits_from_sheet()
        if not self.preview_pending_edits and not self.preview_pending_flags:
            messagebox.showinfo("Hinweis", "Keine Änderungen zum Anwenden.")
            return

        errors, _ = self.entry_service.apply_preview_changes(
            self.preview_pending_edits, self.preview_pending_flags
        )
        if errors:
            messagebox.showerror("Fehler", "\n".join(errors[:10]))
            return

        self.preview_pending_edits = {}
        self.preview_pending_flags = {}
        if self.preview_window is not None and self.preview_window.is_open():
            self.preview_window.set_pending_count(0)
            self.preview_window.clear_pending_highlights()

        self.update_month_view()
        self.update_day_view()
        self.schedule_preview_refresh()

    def reset_preview_changes(self):
        self.preview_pending_edits = {}
        self.preview_pending_flags = {}
        if self.preview_window is not None and self.preview_window.is_open():
            self.preview_window.set_pending_count(0)
            self.preview_window.clear_pending_highlights()
        self.refresh_preview_from_entries(force=True)

    def clear_fields(self, clear_baustelle=True):
        self.entry_hours.delete(0, tk.END)
        self.check_urlaub.set(False)
        self.check_krank.set(False)
        self.check_reise.set(False)
        self.check_fruehstueck.set(False)
        self.check_mittagspause.set(False)
        self.entry_hours.config(state="normal")
        self.entry_hours.delete(0, tk.END)
        self.entry_bst.config(state="normal")

        if clear_baustelle:
            self.entry_bst.delete(0, tk.END)

        self.check_delete_mode.set(False)
        self.clear_edit_mode()

    def get_visible_fields(self):
        return [f for f in self.fields if f.winfo_ismapped()]

    def focus_next(self, event):
        widget = event.widget
        visible_fields = self.get_visible_fields()

        try:
            idx = visible_fields.index(widget)
            next_idx = idx + 1

            if next_idx < len(visible_fields):
                visible_fields[next_idx].focus()
                return "break"
            else:
                self.submit()
                return "break"
        except ValueError:
            pass

    def focus_previous(self, event):
        widget = event.widget
        visible_fields = self.get_visible_fields()

        try:
            idx = visible_fields.index(widget)
            prev_idx = idx - 1

            if prev_idx >= 0:
                visible_fields[prev_idx].focus()
                return "break"
        except ValueError:
            pass

    def setup_autocomplete(self):
        self.name_autocomplete = AutocompleteEntry(
            self.entry_name, self.get_name_suggestions
        )
        self.baustelle_autocomplete = BaustelleAutocomplete(
            self.entry_bst, self.get_baustelle_suggestions
        )

    def get_name_suggestions(self):
        names_data = self.master_db.get_all_names()
        return [n["name"] for n in names_data]

    def get_baustelle_suggestions(self):
        return self.master_db.get_all_baustellen()

    def sort_month_tree(self, col):
        if col == self.month_sort_column:
            self.month_sort_reverse = not self.month_sort_reverse
        else:
            self.month_sort_column = col
            self.month_sort_reverse = False

        items = [
            (self.month_tree.set(item, col), item)
            for item in self.month_tree.get_children("")
        ]

        if col in ("Tag", "Stunden"):
            try:
                items.sort(
                    key=lambda x: float(x[0]) if x[0] else 0,
                    reverse=self.month_sort_reverse,
                )
            except ValueError:
                items.sort(reverse=self.month_sort_reverse)
        else:
            items.sort(reverse=self.month_sort_reverse)

        for index, (val, item) in enumerate(items):
            self.month_tree.move(item, "", index)

        for column in self.month_tree["columns"]:
            heading_text = column
            if column == col:
                heading_text = f"{column} {'▼' if self.month_sort_reverse else '▲'}"
            self.month_tree.heading(column, text=heading_text)

    def sort_day_tree(self, col):
        if col == self.day_sort_column:
            self.day_sort_reverse = not self.day_sort_reverse
        else:
            self.day_sort_column = col
            self.day_sort_reverse = False
        items = [
            (self.day_tree.set(item, col), item)
            for item in self.day_tree.get_children("")
        ]

        if col in ("Tag", "Stunden"):
            try:
                items.sort(
                    key=lambda x: float(x[0]) if x[0] else 0,
                    reverse=self.day_sort_reverse,
                )
            except ValueError:
                items.sort(reverse=self.day_sort_reverse)
        else:
            items.sort(reverse=self.day_sort_reverse)

        for index, (val, item) in enumerate(items):
            self.day_tree.move(item, "", index)

        for column in self.day_tree["columns"]:
            heading_text = column
            if column == col:
                heading_text = f"{column} {'▼' if self.day_sort_reverse else '▲'}"
            self.day_tree.heading(column, text=heading_text)

    def open_name_manager(self):
        NameManagerDialog(self.root)

    def open_baustelle_manager(self):
        BaustelleManagerDialog(self.root)

    def open_settings(self):
        dialog = SettingsDialog(self.root, self.settings, self.master_db)
        self.root.wait_window(dialog.dialog)
        self.settings.current_settings = self.settings.load()

    def get_entry_id_from_tags(self, tags):
        for tag in tags:
            if isinstance(tag, str) and tag.startswith("entry_"):
                return int(tag.split("_")[1])
        return None

    def set_edit_mode(self, entry_id, jahr, monat, tag, name):
        self.edit_mode_active = True
        self.edit_entry_id = entry_id
        self.edit_original_year = str(jahr).strip()
        self.edit_original_month = str(monat).strip()
        self.edit_original_day = str(tag).strip()
        self.edit_original_name = str(name).strip()
        self.update_edit_indicator()

    def clear_edit_mode(self, event=None):
        self.edit_mode_active = False
        self.edit_entry_id = None
        self.edit_original_year = None
        self.edit_original_month = None
        self.edit_original_day = None
        self.edit_original_name = None
        self.update_edit_indicator()

    def load_metadata_flags(self, jahr, monat, tag, name):
        metadata = self.db.get_metadata_by_date(int(jahr), int(monat), int(tag), name)
        if not metadata:
            self.check_fruehstueck.set(0)
            self.check_mittagspause.set(0)
            self.check_skug.set(0)
            self.check_reise.set(0)
            self.combo_reise_type.set(TravelStatus.Nicht)
            self.combo_reise_type.config(state="disabled")
            self.check_urlaub.set(0)
            self.check_krank.set(0)
            self.check_delete_mode.set(False)
            return

        if metadata.get("krank") not in (None, ""):
            self.check_krank.set(1)
            self.toggle_krank()
        elif metadata.get("urlaub") not in (None, ""):
            self.check_urlaub.set(1)
            self.toggle_urlaub()
        else:
            self.check_urlaub.set(0)
            self.check_krank.set(0)

        self.check_fruehstueck.set(1 if metadata.get("fruehstueck") else 0)
        self.check_mittagspause.set(1 if metadata.get("mittag") else 0)
        self.check_skug.set(1 if metadata.get("no_skug") else 0)

        travel_status = metadata.get("travel_status")
        if travel_status:
            self.check_reise.set(1)
            if travel_status in [ts.value for ts in TravelStatus]:
                self.combo_reise_type.set(travel_status)
            else:
                self.combo_reise_type.set(TravelStatus.Nicht)
            self.combo_reise_type.config(state="readonly")
        else:
            self.check_reise.set(0)
            self.combo_reise_type.set(TravelStatus.Nicht)
            self.combo_reise_type.config(state="disabled")

        self.check_delete_mode.set(False)

    def get_urlaub_kostenstelle(self):
        name = self.entry_name.get().strip()
        worker_type = self.master_db.get_worker_type_by_name(name) or WorkerTypes.Fest
        return "900" if worker_type == WorkerTypes.Fest else "940"

    def update_edit_indicator(self):
        if self.edit_mode_active:
            self.edit_status_label.config(text="Bearbeiten: AN", fg="green")
            self.btn_abort_edit.config(state="normal")
        else:
            self.edit_status_label.config(text="Bearbeiten: AUS", fg="gray")
            self.btn_abort_edit.config(state="disabled")

    def check_edit_mode_abort(self, event=None):
        if not self.edit_mode_active:
            return
        if (
            self.entry_year.get().strip() != (self.edit_original_year or "")
            or self.entry_month.get().strip() != (self.edit_original_month or "")
            or self.entry_day.get().strip() != (self.edit_original_day or "")
            or self.entry_name.get().strip() != (self.edit_original_name or "")
        ):
            self.clear_edit_mode()
